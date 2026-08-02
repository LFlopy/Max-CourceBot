
import os
import tempfile
from datetime import datetime
from max_client import MaxBot
import config as cfg
from config import ADMIN_IDS
import database as db
import admin_keyboards as akb

from fsm import set_state, get_state, clear_state, user_states
from utils import (
    build_inline_keyboard,
    build_tariff_deep_link,
    build_user_name,
    build_user_template_context,
    format_inline_buttons_message,
    format_template,
    parse_inline_button_lines,
    parse_duration_to_minutes,
    user_link,
)


def is_admin(user_id: int) -> bool:
    """Return True when the user is configured as an admin."""
    return user_id in ADMIN_IDS


def _parse_duration_to_minutes(text: str) -> int | None:
    return parse_duration_to_minutes(text)


_show_price_in_name = False


def _parse_admin_tariff_id(payload: str, prefix: str) -> int | None:
    if not payload.startswith(prefix):
        return None
    raw_id = payload[len(prefix):].strip()
    if not raw_id.isdigit():
        return None
    tariff_id = int(raw_id)
    return tariff_id if tariff_id > 0 else None


async def handle_admin_callback(bot: MaxBot, update: dict) -> bool:
    """Обрабатывает callback-и с префиксом 'adm:'.
    Возвращает True если обработал, False — если не админский callback.
    """
    callback = update.get("callback", {})
    msg = update.get("message", {})

    callback_id = callback.get("callback_id", "")
    payload = callback.get("payload", "")
    sender = callback.get("user", callback.get("sender", {}))
    user_id = int(sender.get("user_id", 0))

    if not payload.startswith("adm:"):
        return False

    if not is_admin(user_id):
        await bot.answer_callback(callback_id, text="⛔ Нет доступа")
        return True

    body = msg.get("body", {})
    message_id = body.get("mid", "")
    recipient = msg.get("recipient", {})
    chat_id = int(recipient.get("chat_id") or user_id)

    await bot.answer_callback(callback_id)

    async def reply(text: str, keyboard=None):
        ok = await bot.edit_message(message_id, text, keyboard=keyboard)
        if not ok:
            await bot.send_message(chat_id, text, keyboard=keyboard)

    global _show_price_in_name

    if payload == "adm:back" or payload == "adm:menu":
        clear_state(user_id)
        await reply("Административная панель", keyboard=akb.admin_main())

    elif payload == "adm:stats":
        s = await db.stats_summary()
        text = (
            "📊 Статистика\n\n"
            f"👤 Всего пользователей: {s['total_users']}\n"
            f"🆕 Новых за 30 дней: {s['new_users_30d']}\n\n"
            f"🛒 Активных покупок: {s['total_purchases']}\n"
            f"🛒 Покупок за 30 дней: {s['purchases_30d']}\n"
            f"💰 Выручка за 30 дней: {s['revenue_30d']:.0f}₽"
        )
        await reply(text, keyboard=akb.admin_main())

    elif payload == "adm:tariffs":
        clear_state(user_id)
        tariffs = await db.list_tariffs()
        await reply(
            "Список ваших категорий и тарифов",
            keyboard=akb.admin_tariff_list(tariffs, show_price=_show_price_in_name),
        )

    elif payload == "adm:gifts":
        tariffs = await db.list_tariffs()
        selected: set[int] = set()
        set_state(user_id, "adm_gift_pick", selected_tariffs=selected)
        await reply(
            "Укажите тарифы, связанные с гифт файлом.",
            keyboard=akb.admin_gift_tariff_picker(tariffs, selected),
        )

    elif payload.startswith("adm:gift_toggle:"):
        tid = int(payload.split(":")[2])
        state_data = user_states.get(user_id, {})
        selected: set[int] = state_data.get("selected_tariffs", set())
        if tid in selected:
            selected.discard(tid)
        else:
            selected.add(tid)
        state_data["selected_tariffs"] = selected
        tariffs = await db.list_tariffs()
        await reply(
            "Укажите тарифы, связанные с гифт файлом.",
            keyboard=akb.admin_gift_tariff_picker(tariffs, selected),
        )

    elif payload == "adm:gift_next":
        state_data = user_states.get(user_id, {})
        selected: set[int] = state_data.get("selected_tariffs", set())
        set_state(user_id, "adm_gift_wait_file", selected_tariffs=selected)
        await reply("Отправьте гифт файл", keyboard=akb.admin_gift_wait_file())

    elif payload == "adm:toggle_price":
        _show_price_in_name = not _show_price_in_name
        tariffs = await db.list_tariffs()
        await reply(
            "Список ваших категорий и тарифов",
            keyboard=akb.admin_tariff_list(tariffs, show_price=_show_price_in_name),
        )

    elif payload == "adm:reorder":
        set_state(user_id, "adm_reorder")
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф для перемещения",
            keyboard=akb.admin_reorder_list(tariffs),
        )

    elif payload.startswith("adm:sel_reorder:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_reorder", selected_id=tid)
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф для перемещения",
            keyboard=akb.admin_reorder_list(tariffs, selected_id=tid),
        )

    elif payload.startswith("adm:move_up:"):
        tid = int(payload.split(":")[2])
        await db.move_tariff_up(tid)
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф для перемещения",
            keyboard=akb.admin_reorder_list(tariffs, selected_id=tid),
        )

    elif payload.startswith("adm:move_down:"):
        tid = int(payload.split(":")[2])
        await db.move_tariff_down(tid)
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф для перемещения",
            keyboard=akb.admin_reorder_list(tariffs, selected_id=tid),
        )

    elif payload == "adm:add_tariff":
        set_state(user_id, "adm_create_name")
        await reply(
            "Добавление нового тарифа\n \nОтправьте боту название для тарифа:",
            keyboard=akb.admin_create_cancel(),
        )

    elif payload == "adm:add_category":
        set_state(user_id, "adm_create_category")
        await reply(
            "Введите название новой категории:",
            keyboard=akb.admin_create_cancel(),
        )

    elif payload == "adm:cancel_create":
        clear_state(user_id)
        tariffs = await db.list_tariffs()
        await reply(
            "Список ваших категорий и тарифов",
            keyboard=akb.admin_tariff_list(tariffs, show_price=_show_price_in_name),
        )

    elif payload == "adm:create_free":
        state_data = user_states.get(user_id, {})
        name = state_data.get("tariff_name", "")
        set_state(user_id, "adm_create_resources", tariff_name=name, tariff_price=0, is_free=True)
        await reply(
            f"Название тарифа — {name}\n \nЦена: бесплатно\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )

    elif payload.startswith("adm:dur:"):
        dur_val = payload.split(":")[2]
        state_data = user_states.get(user_id, {})
        current_state = get_state(user_id)

        if current_state == "adm_edit_duration":
            tid = state_data.get("tariff_id")
            if dur_val == "custom":
                set_state(user_id, "adm_edit_duration_custom", tariff_id=tid)
                await reply(
                    "Введите срок действия в формате: число часов или минут\n"
                    "Например: 48ч или 120м",
                    keyboard=akb.admin_back_to_settings(tid),
                )
                return True
            days = int(dur_val)
            dur_text = "Навсегда" if days == 0 else f"{days} дн."
            await db.update_tariff(
                tid,
                duration_days=days if days > 0 else 0,
                duration_text=dur_text,
                duration_minutes=None,
                end_date=None,
            )
            clear_state(user_id)
            tariff = await db.get_tariff(tid)
            await bot.send_message(chat_id, "Продолжительность обновлена ✅")
            await _send_tariff_settings(bot, chat_id, tariff)
            return True

        name = state_data.get("tariff_name", "")
        price = state_data.get("tariff_price", 0)
        is_free = state_data.get("is_free", False)

        if dur_val == "custom":
            set_state(user_id, "adm_create_duration_custom",
                      tariff_name=name, tariff_price=price, is_free=is_free)
            await reply(
                "Введите срок действия в формате: число часов или минут\n"
                "Например: 48ч или 120м",
                keyboard=akb.admin_create_cancel(),
            )
            return True

        days = int(dur_val)
        dur_text = "Навсегда" if days == 0 else f"{days} дн."
        set_state(user_id, "adm_create_resources",
                  tariff_name=name, tariff_price=price, is_free=is_free,
                  duration_days=days if days > 0 else None, duration_text=dur_text)

        price_str = "бесплатно" if is_free else f"{price}₽"
        await reply(
            f"Название тарифа — {name}\n \n"
            f"Цена: {price_str}\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, "
            "нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )

    elif payload == "adm:go_resources":
        state_data = user_states.get(user_id, {})
        chats = await db.get_all_bot_chats()
        set_state(user_id, "adm_create_pick_resources",
                  tariff_name=state_data.get("tariff_name", ""),
                  tariff_price=state_data.get("tariff_price", 0),
                  is_free=state_data.get("is_free", False),
                  duration_days=state_data.get("duration_days"),
                  duration_text=state_data.get("duration_text", ""),
                  chats=chats, selected_resources=set())

        name = state_data.get("tariff_name", "")
        price = state_data.get("tariff_price", 0)
        is_free = state_data.get("is_free", False)
        price_str = "бесплатно" if is_free else f"{price}₽"
        await reply(
            f"Тариф — {name}\nЦена: {price_str}\n \n"
            "Выберите ресурсы к которым нужно выдать доступ после покупки",
            keyboard=akb.admin_resource_picker(chats, set()),
        )

    elif payload.startswith("adm:res_pick_page:"):
        page = int(payload.split(":")[2])
        state_data = user_states.get(user_id, {})
        chats = state_data.get("chats", [])
        selected = state_data.get("selected_resources", set())
        is_edit = state_data.get("state") == "adm_edit_resources"
        edit_tid = state_data.get("tariff_id") if is_edit else None
        state_data["resource_page"] = page

        if is_edit:
            tariff = await db.get_tariff(edit_tid)
            header = f"Тариф — {tariff['name']}\n \n"
        else:
            name = state_data.get("tariff_name", "")
            price = state_data.get("tariff_price", 0)
            is_free = state_data.get("is_free", False)
            price_str = "бесплатно" if is_free else f"{price}₽"
            header = f"Тариф — {name}\nЦена: {price_str}\n \n"

        await reply(
            header + "Выберите ресурсы к которым нужно выдать доступ после покупки",
            keyboard=akb.admin_resource_picker(chats, selected, edit_tariff_id=edit_tid, page=page),
        )

    elif payload.startswith("adm:res_toggle:"):
        parts = payload.split(":")
        res_id = int(parts[2])
        page = int(parts[3]) if len(parts) > 3 else 0
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_resources", set())
        if res_id in selected:
            selected.discard(res_id)
        else:
            selected.add(res_id)
        state_data["selected_resources"] = selected
        state_data["resource_page"] = page

        chats = state_data.get("chats", [])
        is_edit = state_data.get("state") == "adm_edit_resources"
        edit_tid = state_data.get("tariff_id") if is_edit else None

        if is_edit:
            tariff = await db.get_tariff(edit_tid)
            header = f"Тариф — {tariff['name']}\n \n"
        else:
            name = state_data.get("tariff_name", "")
            price = state_data.get("tariff_price", 0)
            is_free = state_data.get("is_free", False)
            price_str = "бесплатно" if is_free else f"{price}₽"
            header = f"Тариф — {name}\nЦена: {price_str}\n \n"

        await reply(
            header + "Выберите ресурсы к которым нужно выдать доступ после покупки",
            keyboard=akb.admin_resource_picker(chats, selected, edit_tariff_id=edit_tid, page=page),
        )

    elif payload == "adm:res_save":
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_resources", set())
        chats = state_data.get("chats", [])
        is_edit = state_data.get("state") == "adm_edit_resources"

        resources = []
        for c in chats:
            cid = c.get("chat_id")
            if cid in selected:
                title = c.get("title", "")
                if title.startswith("❓ ") and title.endswith(" (удалён)"):
                    title = title[2:].rsplit(" (удалён)", 1)[0]
                invite_link = c.get("link", "")
                resources.append({"chat_id": cid, "chat_title": title, "invite_link": invite_link})

        if is_edit:
            tid = state_data.get("tariff_id")
            await db.set_tariff_resources(tid, resources)
            clear_state(user_id)
            tariff = await db.get_tariff(tid)
            await reply(
                f"✅ Ресурсы тарифа «{tariff['name']}» обновлены ({len(resources)} шт.)",
                keyboard=akb.admin_tariff_settings(tid, tariff["is_active"]),
            )
        else:
            name = state_data.get("tariff_name", "")
            price = state_data.get("tariff_price", 0)
            is_free = state_data.get("is_free", False)
            duration_days = state_data.get("duration_days")
            duration_minutes = state_data.get("duration_minutes")
            duration_text = state_data.get("duration_text", "")

            tariff = await db.create_tariff(
                name=name, price=float(price), is_free=is_free,
                duration_days=duration_days,
                duration_minutes=duration_minutes,
                duration_text=duration_text,
            )
            if resources:
                await db.set_tariff_resources(tariff["id"], resources)

            clear_state(user_id)
            await reply(
                f"Тариф «{name}»\nУспешно создан ✅",
                keyboard=akb.admin_tariff_created(tariff["id"]),
            )

    elif payload.startswith("adm:tariff:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        if not tariff:
            await reply("Тариф не найден")
            return True
        await _show_tariff_settings(reply, tariff)

    elif payload.startswith("adm:settings:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        if not tariff:
            await reply("Тариф не найден")
            return True
        await _show_tariff_settings(reply, tariff)

    elif payload.startswith("adm:save_settings:"):
        tid = int(payload.split(":")[2])
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await reply(
            f"Настройки тарифа «{tariff['name']}» сохранены ✅",
            keyboard=akb.admin_tariff_settings(tid, tariff["is_active"]),
        )

    elif payload.startswith("adm:toggle_active:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        new_active = not tariff["is_active"]
        await db.update_tariff(tid, is_active=new_active)
        tariff = await db.get_tariff(tid)
        await _show_tariff_settings(reply, tariff)

    elif payload.startswith("adm:delete:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        await reply(
            f"⚠️ Вы уверены что хотите удалить тариф «{tariff['name']}»?",
            keyboard=akb.admin_confirm_delete(tid),
        )

    elif payload.startswith("adm:confirm_del:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        name = tariff["name"] if tariff else "?"
        await db.delete_tariff(tid)
        tariffs = await db.list_tariffs()
        await reply(
            f"Тариф «{name}» удалён 🗑",
            keyboard=akb.admin_tariff_list(tariffs, show_price=_show_price_in_name),
        )

    elif payload.startswith("adm:set_name:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_name", tariff_id=tid)
        await reply(
            "Отправьте боту новое название тарифа:",
            keyboard=akb.admin_back_to_settings(tid),
        )

    elif payload.startswith("adm:set_price:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        t = tariff
        await reply(
            f"Настройка цен тарифа «{t['name']}»\n \n"
            f"Стандартная цена: {t['price']}₽\n"
            f"Старая цена: {t['old_price'] or '-'}\n"
            f"Цена продления: {t['renewal_price'] or '-'}\n"
            f"Цена продления активной подписки: {t['active_renewal_price'] or '-'}",
            keyboard=akb.admin_price_settings(tid),
        )

    elif payload.startswith("adm:ep_std:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_price_std", tariff_id=tid)
        await reply("Введите новую стандартную цену:", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:ep_old:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_price_old", tariff_id=tid)
        await reply("Введите новую старую цену (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:ep_renew:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_price_renew", tariff_id=tid)
        await reply("Введите цену продления (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:ep_active:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_price_active", tariff_id=tid)
        await reply("Введите цену продления активной подписки (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_desc:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_desc", tariff_id=tid)
        await reply(
            "Отправьте боту новое описание тарифа\n(Или delete, чтобы удалить его):",
            keyboard=akb.admin_back_to_settings(tid),
        )

    elif payload.startswith("adm:set_duration:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_duration", tariff_id=tid)
        tariff = await db.get_tariff(tid)
        await reply(
            f"Текущая продолжительность: {tariff['duration_text'] or tariff['duration_days'] or 'не задана'}\n \n"
            "Введите новый срок действия тарифа в днях "
            "или выберите из готовых вариантов ниже",
            keyboard=akb.admin_edit_duration(tid),
        )

    elif payload.startswith("adm:set_dates:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        start_day = tariff["start_day"] or "не установлен"
        start_date = tariff["start_date"].strftime("%d.%m.%Y %H:%M") if tariff["start_date"] else "не установлен"
        end_date = tariff["end_date"].strftime("%d.%m.%Y %H:%M") if tariff["end_date"] else "не установлен"
        await reply(
            f"Тариф: {tariff['name']}\n \n"
            f"Дата начала тарифа: {start_date}\n"
            f"Дата конца тарифа: {end_date}",
            keyboard=akb.admin_date_settings(tid),
        )


    elif payload.startswith("adm:ed_start:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_start_date", tariff_id=tid)
        await reply(
            'Введите дату (в формате: 16.03.2026 07:54) или "/empty" чтобы сбросить её',
            keyboard=akb.admin_date_input_back(tid),
        )

    elif payload.startswith("adm:ed_end:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_end_date", tariff_id=tid)
        await reply(
            'Введите дату (в формате: 16.03.2026 07:54) или "/empty" чтобы сбросить её',
            keyboard=akb.admin_date_input_back(tid),
        )

    elif payload.startswith("adm:set_resources:"):
        tid = int(payload.split(":")[2])
        chats = await db.get_all_bot_chats()
        existing = await db.get_tariff_resources(tid)
        selected = {r["chat_id"] for r in existing}
        existing_links = {r["chat_id"]: r.get("invite_link", "") for r in existing}
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        for c in chats:
            cid = c.get("chat_id")
            if not c.get("link") and cid in existing_links and existing_links[cid]:
                c["link"] = existing_links[cid]
        for r in existing:
            if r["chat_id"] not in chat_ids_in_list:
                chats.append({
                    "chat_id": r["chat_id"],
                    "title": f"❓ {r.get('chat_title') or r['chat_id']} (удалён)",
                    "link": r.get("invite_link", ""),
                })
        set_state(user_id, "adm_edit_resources", tariff_id=tid, chats=chats, selected_resources=selected, resource_page=0)
        tariff = await db.get_tariff(tid)
        await reply(
            f"Тариф — {tariff['name']}\n \n"
            "Выберите ресурсы к которым нужно выдать доступ после покупки",
            keyboard=akb.admin_resource_picker(chats, selected, edit_tariff_id=tid),
        )

    elif payload.startswith("adm:set_check:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_check_name", tariff_id=tid)
        await reply("Введите название в чеке (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_reject:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_reject", tariff_id=tid)
        await reply("Введите интервал отклонений в минутах (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_success:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_success", tariff_id=tid)
        await reply("Введите текст при успешной покупке (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_limit:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_limit", tariff_id=tid)
        await reply("Введите лимит активаций (число, или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_allowed:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        tariffs = await db.list_tariffs()
        selected = set()
        if tariff.get("allowed_group"):
            selected = {int(x) for x in tariff["allowed_group"].split(",") if x.strip()}
        set_state(user_id, "adm_allowed_pick", tariff_id=tid, selected_allowed=selected)
        await reply(
            f"Тариф «{tariff['name']}»\n\n"
            "Выберите тарифы, подписчикам которых будет виден этот тариф.\n"
            "Если ни один не выбран — тариф виден всем.",
            keyboard=akb.admin_tariff_allowed_picker(tariffs, selected, tid),
        )

    elif payload.startswith("adm:allowed_toggle:"):
        parts = payload.split(":")
        tid = int(parts[2])
        toggle_tid = int(parts[3])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_allowed", set())
        if toggle_tid in selected:
            selected.discard(toggle_tid)
        else:
            selected.add(toggle_tid)
        state_data["selected_allowed"] = selected
        tariffs = await db.list_tariffs()
        tariff = await db.get_tariff(tid)
        await reply(
            f"Тариф «{tariff['name']}»\n\n"
            "Выберите тарифы, подписчикам которых будет виден этот тариф.\n"
            "Если ни один не выбран — тариф виден всем.",
            keyboard=akb.admin_tariff_allowed_picker(tariffs, selected, tid),
        )

    elif payload.startswith("adm:allowed_save:"):
        tid = int(payload.split(":")[2])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_allowed", set())
        val = ",".join(str(x) for x in selected) if selected else None
        await db.update_tariff(tid, allowed_group=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await reply("Группа разрешённых обновлена ✅")
        await _show_tariff_settings(reply, tariff)

    elif payload.startswith("adm:tariff_link:"):
        tid = _parse_admin_tariff_id(payload, "adm:tariff_link:")
        if tid is None:
            await bot.send_message(chat_id, "❌ Некорректный ID тарифа.")
            return True

        tariff = await db.get_tariff(tid)
        if not tariff:
            await bot.send_message(chat_id, "❌ Тариф не найден. Возможно, он был удалён.")
            return True

        try:
            link = build_tariff_deep_link(getattr(cfg, "MAX_BOT_USERNAME", ""), tid)
        except ValueError:
            await bot.send_message(
                chat_id,
                "❌ Не задан MAX_BOT_USERNAME.\n\n"
                "Укажите публичное имя бота в config.py и перезапустите приложение.",
            )
            return True

        await bot.send_message(
            chat_id,
            f"🔗 Ссылка на тариф «{tariff['name']}»\n\n"
            f"{link}\n\n"
            "Ссылку можно использовать в рассылках, публикациях и рекламных материалах.\n"
            "Пользователь, перешедший по ней, попадёт сразу к оформлению этого тарифа.",
        )
        await db.add_user_log(user_id, f"Получил deep link тарифа «{tariff['name']}»")

    elif payload.startswith("adm:buy_link:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        if tariff["is_free"]:
            btns = [
                [{"type": "callback", "text": "✅ Активировать тариф", "payload": f"activate:{tid}"}],
                [{"type": "callback", "text": "🔙 Назад", "payload": f"adm:settings:{tid}"}],
            ]
            await reply(
                f"Тариф «{tariff['name']}» — бесплатный.\n\n"
                "Нажмите кнопку ниже, чтобы активировать:",
                keyboard={"type": "inline_keyboard", "payload": {"buttons": btns}},
            )
        else:
            link = tariff.get("payment_link") or ""
            btns = []
            if link:
                btns.append([{"type": "link", "text": "💳 Перейти к оплате", "url": link}])
            btns.append([{"type": "callback", "text": "💳 Вызвать оплату в боте", "payload": f"pay:{tid}"}])
            btns.append([{"type": "callback", "text": "🔙 Назад", "payload": f"adm:settings:{tid}"}])
            text = f"Тариф «{tariff['name']}» — {tariff['price']}₽\n\n"
            if link:
                text += f"Ссылка на оплату: {link}\n\n"
            text += "Нажмите кнопку ниже для перехода к покупке:"
            await reply(
                text,
                keyboard={"type": "inline_keyboard", "payload": {"buttons": btns}},
            )

    elif payload.startswith("adm:set_category:"):
        tid = int(payload.split(":")[2])
        categories = await db.list_categories()
        tariff = await db.get_tariff(tid)
        btns = []
        for cat in categories:
            icon = "✅" if tariff["category_id"] == cat["id"] else "⬜"
            btns.append([{"type": "callback", "text": f"{icon} {cat['name']}", "payload": f"adm:pick_cat:{tid}:{cat['id']}"}])
        btns.append([{"type": "callback", "text": "❌ Без категории", "payload": f"adm:pick_cat:{tid}:0"}])
        btns.append([{"type": "callback", "text": "🔙 Назад", "payload": f"adm:settings:{tid}"}])
        await reply(
            f"Выберите категорию для тарифа «{tariff['name']}»:",
            keyboard={"type": "inline_keyboard", "payload": {"buttons": btns}},
        )

    elif payload.startswith("adm:pick_cat:"):
        parts = payload.split(":")
        tid = int(parts[2])
        cat_id = int(parts[3])
        await db.update_tariff(tid, category_id=cat_id if cat_id > 0 else None)
        tariff = await db.get_tariff(tid)
        await _show_tariff_settings(reply, tariff)

    elif payload.startswith("adm:set_media:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_media", tariff_id=tid)
        await reply("Отправьте ссылку на медиа (изображение) для тарифа\n(Или /empty чтобы удалить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:tariff_gifts:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        if not tariff:
            await reply("Тариф не найден")
            return True
        gifts = await db.get_gift_files_for_tariff(tid)
        lines = []
        for g in gifts:
            name = g.get("file_name") or f"Файл #{g['id']}"
            lines.append(f"• {name}")
        if lines:
            text = "🎁 Бонусные файлы тарифа «" + tariff["name"] + "»:\n\n" + "\n".join(lines)
        else:
            text = "🎁 Бонусные файлы тарифа «" + tariff["name"] + "»:\n\nФайлов нет."
        await reply(text, keyboard=akb.admin_tariff_gifts_menu(tid))

    elif payload.startswith("adm:tariff_gift_add:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_tariff_gift_wait_file", tariff_id=tid)
        await reply("Отправьте боту новый бонусный файл:", keyboard=akb.admin_tariff_gift_wait_file(tid))

    elif payload.startswith("adm:tariff_gift_del:"):
        tid = int(payload.split(":")[2])
        gifts = await db.get_gift_files_for_tariff(tid)
        if not gifts:
            await reply("Бонусных файлов нет.", keyboard=akb.admin_tariff_gifts_menu(tid))
            return True
        await reply("Выберите бонусный файл, который хотите удалить:", keyboard=akb.admin_tariff_gift_delete_list(gifts, tid))

    elif payload.startswith("adm:tariff_gift_del_confirm:"):
        parts = payload.split(":")
        gift_id = int(parts[2])
        tid = int(parts[3])
        await db.delete_gift_file(gift_id)
        gifts = await db.get_gift_files_for_tariff(tid)
        lines = []
        for g in gifts:
            name = g.get("file_name") or f"Файл #{g['id']}"
            lines.append(f"• {name}")
        tariff = await db.get_tariff(tid)
        if lines:
            text = "✅ Файл удалён.\n\n🎁 Бонусные файлы тарифа «" + tariff["name"] + "»:\n\n" + "\n".join(lines)
        else:
            text = "✅ Файл удалён.\n\n🎁 Бонусные файлы тарифа «" + tariff["name"] + "»:\n\nФайлов нет."
        await reply(text, keyboard=akb.admin_tariff_gifts_menu(tid))

    elif payload.startswith("adm:warmup_list:"):
        tid = int(payload.split(":")[2])
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        if not tariff:
            await reply("Тариф не найден")
            return True
        messages = await db.get_warmup_messages(tid)
        await reply(
            f"🔥 Догревающие рассылки тарифа «{tariff['name']}»\n\n"
            "Отправляются пользователям, которые начали оплату, но не завершили её.",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
            )

    elif payload.startswith("adm:warmup_toggle_mode:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        current_mode = tariff.get("warmup_order_mode") or "sequential"
        new_mode = "random" if current_mode == "sequential" else "sequential"
        await db.set_tariff_warmup_order_mode(tid, new_mode)
        messages = await db.get_warmup_messages(tid)
        await reply(
            f"🔥 Догревающие рассылки тарифа «{tariff['name']}»\n\n"
            "Отправляются пользователям, которые начали оплату, но не завершили её.",
            keyboard=akb.admin_warmup_list(tid, messages, new_mode),
        )

    elif payload.startswith("adm:warmup_add:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_warmup_add_text", tariff_id=tid)
        await reply(
            "Отправьте текст догревающего сообщения:",
            keyboard=akb.admin_warmup_cancel(tid),
        )

    elif payload.startswith("adm:warmup_open:"):
        mid = int(payload.split(":")[2])
        msg = await db.get_warmup_message(mid)
        if not msg:
            await reply("Сообщение не найдено")
            return True
        delay = msg["delay_minutes"]
        delay_label = f"{delay} мин."
        if delay >= 1440:
            delay_label = f"{delay // 1440} дн."
        elif delay >= 60:
            delay_label = f"{delay // 60} ч."
        await reply(
            f"Текст: {msg['text']}\n\n"
            f"Медиа: {'есть' if msg['media_url'] else 'нет'}\n"
            f"Кнопки: {len(msg.get('buttons') or [])}\n"
            f"Время отправки: через {delay_label} после начала оплаты\n"
            f"Статус: {'активно' if msg['is_active'] else 'отключено'}",
            keyboard=akb.admin_warmup_detail(mid, msg["tariff_id"], msg["is_active"]),
        )

    elif payload.startswith("adm:warmup_buttons_menu:"):
        parts = payload.split(":")
        mid, tid = int(parts[2]), int(parts[3])
        msg = await db.get_warmup_message(mid)
        buttons = msg.get("buttons") or []
        if buttons:
            await reply(
                format_inline_buttons_message(buttons),
                keyboard=akb.admin_warmup_button_list(mid, tid, buttons, len(buttons) < 5),
            )
        else:
            await reply(
                "➕ Добавить кнопки к догревающей рассылке?\n\n"
                "Можно комбинировать тарифы и сторонние ссылки (до 5 кнопок).",
                keyboard=akb.admin_warmup_button_list(mid, tid, buttons, True),
            )

    elif payload.startswith("adm:warmup_buttons_clear:"):
        parts = payload.split(":")
        mid, tid = int(parts[2]), int(parts[3])
        await db.update_warmup_message(mid, buttons=[])
        await reply(
            "✅ Кнопки удалены.",
            keyboard=akb.admin_warmup_back_to_detail(mid, tid),
        )

    elif payload.startswith("adm:warmup_add_btn:"):
        parts = payload.split(":")
        if len(parts) == 4:
            mid, tid = int(parts[2]), int(parts[3])
            msg = await db.get_warmup_message(mid)
            set_state(user_id, "adm_warmup_wait_input_button",
                      message_id=mid, tariff_id=tid, warmup_buttons=msg.get("buttons") or [])
        else:
            tid = int(parts[2])
            sd = user_states.get(user_id, {})
            set_state(user_id, "adm_warmup_add_wait_input_button",
                      tariff_id=tid,
                      warmup_text=sd.get("warmup_text"),
                      warmup_delay_minutes=sd.get("warmup_delay_minutes"),
                      warmup_buttons=sd.get("warmup_buttons", []))
        await reply(
            "📌 Отправьте кнопку в формате:\n"
            "**Текст кнопки - https://ссылка.com**\n"
            "Подойдут `-`, `--` и `—`.",
            keyboard=akb.admin_warmup_cancel(tid),
        )

    elif payload.startswith("adm:warmup_add_btns:tariff:"):
        parts = payload.split(":")
        if len(parts) == 5:
            mid, tid = int(parts[3]), int(parts[4])
            msg = await db.get_warmup_message(mid)
            buttons = msg.get("buttons") or []
        else:
            mid = None
            tid = int(parts[3])
            sd = user_states.get(user_id, {})
            buttons = sd.get("warmup_buttons", [])
        if len(buttons) >= 5:
            await bot.answer_callback(callback_id, text="Максимум 5 кнопок ⚠️")
            return True
        added_tariff_ids = {
            b["tariff_id"] for b in buttons
            if b.get("kind") == "tariff" and b.get("tariff_id")
        }
        tariffs = await db.list_tariffs()
        await reply(
            "💰 Выберите тариф для кнопки в догревающей рассылке:",
            keyboard=akb.admin_warmup_button_picker(mid, tid, tariffs, added_tariff_ids),
        )

    elif payload == "adm:warmup_add_btn_disabled":
        await bot.answer_callback(callback_id, text="Максимум 5 кнопок достигнут ⚠️")

    elif payload.startswith("adm:warmup_btn_tariff:"):
        parts = payload.split(":")
        mid, warmup_tid, button_tid = int(parts[2]), int(parts[3]), int(parts[4])
        tariff = await db.get_tariff(button_tid)
        if not tariff:
            await bot.answer_callback(callback_id, text="Тариф не найден")
            return True
        if mid:
            msg = await db.get_warmup_message(mid)
            buttons = list(msg.get("buttons") or [])
        else:
            sd = user_states.get(user_id, {})
            buttons = list(sd.get("warmup_buttons", []))
        if len(buttons) >= 5:
            await bot.answer_callback(callback_id, text="Максимум 5 кнопок ⚠️")
            return True
        if any(b.get("kind") == "tariff" and b.get("tariff_id") == button_tid for b in buttons):
            await bot.answer_callback(callback_id, text="Этот тариф уже добавлен")
            return True
        buttons.append({"kind": "tariff", "text": tariff["name"], "tariff_id": button_tid})
        if mid:
            await db.update_warmup_message(mid, buttons=buttons)
            await reply(
                format_inline_buttons_message(buttons),
                keyboard=akb.admin_warmup_button_list(mid, warmup_tid, buttons, len(buttons) < 5),
            )
        else:
            sd = user_states.get(user_id, {})
            set_state(user_id, "adm_warmup_add_buttons_added",
                      tariff_id=warmup_tid,
                      warmup_text=sd.get("warmup_text"),
                      warmup_delay_minutes=sd.get("warmup_delay_minutes"),
                      warmup_buttons=buttons)
            await reply(
                format_inline_buttons_message(buttons),
                keyboard=akb.admin_warmup_button_list(None, warmup_tid, buttons, len(buttons) < 5),
            )

    elif payload.startswith("adm:warmup_add_buttons_menu:"):
        tid = int(payload.split(":")[2])
        sd = user_states.get(user_id, {})
        buttons = sd.get("warmup_buttons", [])
        if buttons:
            await reply(
                format_inline_buttons_message(buttons),
                keyboard=akb.admin_warmup_button_list(None, tid, buttons, len(buttons) < 5),
            )
        else:
            await reply(
                "➕ Добавить кнопки к догревающей рассылке?\n\n"
                "Можно комбинировать тарифы и сторонние ссылки (до 5 кнопок).",
                keyboard=akb.admin_warmup_add_buttons(tid),
            )

    elif payload.startswith("adm:warmup_add_btns:yes:"):
        tid = int(payload.split(":")[3])
        sd = user_states.get(user_id, {})
        set_state(user_id, "adm_warmup_add_wait_input_button",
                  tariff_id=tid,
                  warmup_text=sd.get("warmup_text"),
                  warmup_delay_minutes=sd.get("warmup_delay_minutes"),
                  warmup_buttons=sd.get("warmup_buttons", []))
        await reply(
            "📌 Отправьте кнопки в формате:\n"
            "**Текст кнопки - https://ссылка.com**\n\n"
            "Пример:\n"
            "Статья 1 - https://example.com/article-1\n\n"
            "Подойдут `-`, `--` и `—`.\n"
            "Можно отправлять по одной кнопке или весь список сразу (разделяя строками)",
            keyboard=akb.admin_warmup_cancel(tid),
        )

    elif payload.startswith("adm:warmup_add_btns:no:"):
        tid = int(payload.split(":")[3])
        sd = user_states.get(user_id, {})
        await db.create_warmup_message(
            tid,
            text=sd.get("warmup_text", ""),
            media_url=None,
            delay_minutes=sd.get("warmup_delay_minutes", 0),
            buttons=[],
        )
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        messages = await db.get_warmup_messages(tid)
        await reply(
            f"✅ Сообщение добавлено.\n\n🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )

    elif payload.startswith("adm:warmup_add_save:"):
        tid = int(payload.split(":")[2])
        sd = user_states.get(user_id, {})
        await db.create_warmup_message(
            tid,
            text=sd.get("warmup_text", ""),
            media_url=None,
            delay_minutes=sd.get("warmup_delay_minutes", 0),
            buttons=sd.get("warmup_buttons", []),
        )
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        messages = await db.get_warmup_messages(tid)
        await reply(
            f"✅ Сообщение добавлено.\n\n🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )

    elif payload.startswith("adm:warmup_edit_text:"):
        mid = int(payload.split(":")[2])
        msg = await db.get_warmup_message(mid)
        set_state(user_id, "adm_warmup_edit_text", message_id=mid, tariff_id=msg["tariff_id"])
        await reply(
            "Отправьте новый текст сообщения:",
            keyboard=akb.admin_warmup_back_to_detail(mid, msg["tariff_id"]),
        )

    elif payload.startswith("adm:warmup_edit_media:"):
        mid = int(payload.split(":")[2])
        msg = await db.get_warmup_message(mid)
        set_state(user_id, "adm_warmup_edit_media", message_id=mid, tariff_id=msg["tariff_id"])
        await reply(
            "Отправьте новое медиа (фото/видео) для сообщения\n(Или /empty чтобы убрать медиа):",
            keyboard=akb.admin_warmup_back_to_detail(mid, msg["tariff_id"]),
            )

    elif payload.startswith("adm:warmup_edit_delay:"):
        mid = int(payload.split(":")[2])
        msg = await db.get_warmup_message(mid)
        set_state(user_id, "adm_warmup_edit_delay", message_id=mid, tariff_id=msg["tariff_id"])
        await reply(
            "Введите время отправки после начала оплаты.\n"
            "Формат: число + ч/м, например: 2ч, 30м, 48ч",
            keyboard=akb.admin_warmup_back_to_detail(mid, msg["tariff_id"]),
        )

    elif payload.startswith("adm:warmup_up:"):
        parts = payload.split(":")
        mid, tid = int(parts[2]), int(parts[3])
        await db.move_warmup_message_up(mid)
        messages = await db.get_warmup_messages(tid)
        tariff = await db.get_tariff(tid)
        await reply(
            f"🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )

    elif payload.startswith("adm:warmup_down:"):
        parts = payload.split(":")
        mid, tid = int(parts[2]), int(parts[3])
        await db.move_warmup_message_down(mid)
        messages = await db.get_warmup_messages(tid)
        tariff = await db.get_tariff(tid)
        await reply(
            f"🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )

    elif payload.startswith("adm:warmup_toggle_active:"):
        parts = payload.split(":")
        mid, tid = int(parts[2]), int(parts[3])
        msg = await db.get_warmup_message(mid)
        await db.update_warmup_message(mid, is_active=not msg["is_active"])
        messages = await db.get_warmup_messages(tid)
        tariff = await db.get_tariff(tid)
        await reply(
            f"🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )

    elif payload.startswith("adm:warmup_delete:"):
        parts = payload.split(":")
        mid, tid = int(parts[2]), int(parts[3])
        await db.delete_warmup_message(mid)
        messages = await db.get_warmup_messages(tid)
        tariff = await db.get_tariff(tid)
        await reply(
            f"✅ Сообщение удалено.\n\n🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )


    elif payload == "adm:settings_menu":
        clear_state(user_id)
        await reply(
            "Настройки вашего бота.\n"
            "Для того чтобы вернуться обратно в Главное меню, "
            "вы можете отправить команду /start",
            keyboard=akb.admin_bot_settings(),
        )

    elif payload == "adm:manage_resources":
        clear_state(user_id)
        chats = await db.get_all_bot_chats()
        usage = await db.get_resource_usage()
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        for cid, tariff_names in usage.items():
            if cid not in chat_ids_in_list:
                chats.append({"chat_id": cid, "title": f"❓ {cid} (недоступен)"})
        await reply(
            "📋 Ресурсы бота (чаты/каналы).\n\n"
            "Нажмите на ресурс, чтобы открыть подтверждение удаления.",
            keyboard=akb.admin_resources_list(chats, usage),
        )

    elif payload.startswith("adm:manage_resources_page:"):
        page = int(payload.split(":")[2])
        chats = await db.get_all_bot_chats()
        usage = await db.get_resource_usage()
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        for cid, tariff_names in usage.items():
            if cid not in chat_ids_in_list:
                chats.append({"chat_id": cid, "title": f"❓ {cid} (недоступен)"})
        await reply(
            "📋 Ресурсы бота (чаты/каналы).\n\n"
            "Нажмите на ресурс, чтобы открыть подтверждение удаления.",
            keyboard=akb.admin_resources_list(chats, usage, page=page),
        )

    elif payload == "adm:res_add_manual":
        set_state(user_id, "adm_add_resource")
        await reply(
            "Отправьте `chat_id` чата или канала, где бот уже добавлен.",
            keyboard=akb.admin_back_to_settings_menu(),
        )

    elif payload.startswith("adm:res_delete:"):
        cid = int(payload.split(":")[2])
        chats = await db.get_all_bot_chats()
        usage = await db.get_resource_usage()
        title = str(cid)
        for c in chats:
            if c.get("chat_id") == cid:
                title = c.get("title", str(cid))
                break
        tariff_names = usage.get(cid, [])
        text = f"Удалить ресурс **{title}**?\n\n"
        if tariff_names:
            text += "Используется в тарифах:\n"
            text += "\n".join(f" • {n}" for n in tariff_names)
            text += "\n\nРесурс будет отвязан от всех тарифов."
        else:
            text += "Не привязан ни к одному тарифу."
        text += "\nБот покинет этот чат."
        await reply(text, keyboard=akb.admin_confirm_res_delete(cid))

    elif payload.startswith("adm:res_delete_confirm:"):
        cid = int(payload.split(":")[2])
        chats = await db.get_all_bot_chats()
        title = str(cid)
        for c in chats:
            if c.get("chat_id") == cid:
                title = c.get("title", str(cid))
                break
        left = await bot.leave_chat(cid)
        await db.delete_resource_from_all_tariffs(cid)
        if left:
            await db.mark_bot_chat_removed(cid)

        chats = await db.get_all_bot_chats()
        usage = await db.get_resource_usage()
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        for c_id, _ in usage.items():
            if c_id not in chat_ids_in_list:
                chats.append({"chat_id": c_id, "title": f"❓ {c_id} (недоступен)"})

        if left:
            text = f"✅ Бот вышел из ресурса «{title}». Ресурс отвязан от тарифов.\n\n"
        else:
            text = (
                f"⚠️ Не удалось выполнить выход из ресурса «{title}» через MAX API. "
                "Ресурс отвязан от тарифов, но может оставаться в каталоге до получения события удаления.\n\n"
            )
        await reply(
            text + "📋 Ресурсы бота (чаты/каналы).",
            keyboard=akb.admin_resources_list(chats, usage),
        )

    elif payload.startswith("adm:res_del:"):
        cid = int(payload.split(":")[2])
        chats = await db.get_all_bot_chats()
        usage = await db.get_resource_usage()
        title = str(cid)
        for c in chats:
            if c.get("chat_id") == cid:
                title = c.get("title", str(cid))
                break
        tariff_names = usage.get(cid, [])
        text = f"Удалить ресурс **{title}**?\n\n"
        if tariff_names:
            text += "Используется в тарифах:\n"
            text += "\n".join(f" • {n}" for n in tariff_names)
            text += "\n\nРесурс будет отвязан от всех тарифов."
        else:
            text += "Не привязан ни к одному тарифу."
        text += "\nБот покинет этот чат."
        await reply(text, keyboard=akb.admin_confirm_res_delete(cid))

    elif payload.startswith("adm:res_del_confirm:"):
        cid = int(payload.split(":")[2])
        await db.delete_resource_from_all_tariffs(cid)
        left = await bot.leave_chat(cid)
        if left:
            await db.mark_bot_chat_removed(cid)
        chats = await db.get_all_bot_chats()
        usage = await db.get_resource_usage()
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        for c_id, _ in usage.items():
            if c_id not in chat_ids_in_list:
                chats.append({"chat_id": c_id, "title": f"❓ {c_id} (недоступен)"})
        await reply(
            "✅ Ресурс удалён.\n\n"
            "📋 Ресурсы бота (чаты/каналы).",
            keyboard=akb.admin_resources_list(chats, usage),
        )

    elif payload == "adm:subscribers":
        clear_state(user_id)
        s = await db.subscribers_stats()
        text = (
            f"Всего переходов в бота: **{s['total_users']}** чел.\n\n"
            "Количество подписчиков:\n"
            f" • Активных подписок: **{s['active_subs']}** чел.\n"
            f" • Не продлили подписку: **{s['expired']}** чел.\n"
            f" • Купили подписку: **{s['bought']}** чел.\n"
            f" • Ни разу не купивших: **{s['never_bought']}** чел."
        )
        await reply(text, keyboard=akb.admin_subscribers())

    elif payload == "adm:sub_search_id":
        set_state(user_id, "adm_search_id_input")
        await reply(
            "Введите ID пользователя:",
            keyboard=akb.admin_back_subscribers(),
        )

    elif payload == "adm:sub_profile":
        set_state(user_id, "adm_profile_input")
        await reply(
            "Введите id пользователя, чтобы получить их профиль:",
            keyboard=akb.admin_back_subscribers(),
        )

    elif payload == "adm:sub_list":
        try:
            await _send_subscribers_xlsx(bot, chat_id)
            await bot.send_message(chat_id, "👆 Список подписчиков", keyboard=akb.admin_subscribers())
        except Exception as e:
            await bot.send_message(chat_id, f"❌ Ошибка: {e}", keyboard=akb.admin_subscribers())

    elif payload == "adm:sub_expired":
        try:
            await _send_expired_xlsx(bot, chat_id)
            await bot.send_message(chat_id, "👆 Список не продливших", keyboard=akb.admin_subscribers())
        except Exception as e:
            await bot.send_message(chat_id, f"❌ Ошибка: {e}", keyboard=akb.admin_subscribers())

    elif payload == "adm:sub_table":
        try:
            await _send_users_xlsx(bot, chat_id)
            await bot.send_message(chat_id, "👆 Таблица пользователей", keyboard=akb.admin_subscribers())
        except Exception as e:
            await bot.send_message(chat_id, f"❌ Ошибка: {e}", keyboard=akb.admin_subscribers())

    elif payload == "adm:sub_grant":
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите какую подписку вы хотите подарить пользователю",
            keyboard=akb.admin_grant_tariff_list(tariffs),
        )

    elif payload.startswith("adm:grant_pick:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        set_state(user_id, "adm_grant_user", tariff_id=tid)
        await reply(
            f"Отправьте user\\_ID пользователя которому вы хотите "
            f"выдать подписку «{tariff['name']}».\n\n"
            "**!Обратите внимание, что для выдачи подписки, "
            "пользователь, которому активируем подписку, "
            "должен запустить бота, написав /start**",
            keyboard=akb.admin_back_subscribers(),
        )

    elif payload == "adm:sub_revoke":
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите подписку для обнуления",
            keyboard=akb.admin_revoke_tariff_list(tariffs),
        )

    elif payload.startswith("adm:revoke_pick:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        set_state(user_id, "adm_revoke_user", tariff_id=tid)
        await reply(
            f"Отправьте user\\_ID пользователя, у которого нужно "
            f"обнулить подписку «{tariff['name']}».",
            keyboard=akb.admin_back_subscribers(),
        )

    elif payload == "adm:sub_transfer":
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите подписку для передачи",
            keyboard=akb.admin_transfer_tariff_list(tariffs),
        )

    elif payload.startswith("adm:transfer_pick:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        set_state(user_id, "adm_transfer_from", tariff_id=tid)
        await reply(
            f"Отправьте user\\_ID пользователя, **от которого** "
            f"передать подписку «{tariff['name']}».",
            keyboard=akb.admin_back_subscribers(),
        )

    elif payload.startswith("adm:msg_user:"):
        target = int(payload.split(":")[2])
        set_state(user_id, "adm_msg_user", target_user_id=target)
        await reply(
            f"Введите сообщение для пользователя (id={target}):",
            keyboard=akb.admin_back_subscribers(),
        )

    elif payload.startswith("adm:pay_table:"):
        target = int(payload.split(":")[2])
        profile = await db.user_profile(target)
        if not profile:
            await reply("Пользователь не найден.")
            return
        purchases = profile["purchases"]
        if not purchases:
            await reply("У пользователя нет платежей.",
                        keyboard=akb.admin_user_profile(target))
            return
        lines = [f"🧾 Платежи пользователя {target}:\n"]
        for p in purchases:
            date_str = p["purchased_at"].strftime("%d.%m.%Y %H:%M") if p.get("purchased_at") else "?"
            status_map = {"active": "✅", "pending": "⏳", "expired": "⌛", "canceled": "❌", "revoked": "🚫"}
            icon = status_map.get(p["status"], "❓")
            price = f"{p['price_paid']}₽" if p.get("price_paid") else "бесплатно"
            lines.append(f"{icon} {p['tariff_name']} — {price} — {date_str} ({p['status']})")
        await reply("\n".join(lines), keyboard=akb.admin_user_profile(target))

    elif payload.startswith("adm:sub_table_user:"):
        target = int(payload.split(":")[2])
        profile = await db.user_profile(target)
        if not profile:
            await reply("Пользователь не найден.")
            return
        purchases = profile["purchases"]
        active = [p for p in purchases if p["status"] == "active"]
        if not active:
            await reply("У пользователя нет активных подписок.",
                        keyboard=akb.admin_user_profile(target))
            return
        lines = [f"📋 Подписки пользователя {target}:\n"]
        for p in active:
            exp = p.get("expires_at")
            if exp:
                exp_str = exp.strftime("%d.%m.%Y %H:%M")
            else:
                exp_str = "бессрочно"
            activated = p["activated_at"].strftime("%d.%m.%Y") if p.get("activated_at") else "?"
            lines.append(f"• {p['tariff_name']} — с {activated} до {exp_str}")
        await reply("\n".join(lines), keyboard=akb.admin_user_profile(target))


    elif payload == "adm:broadcast":
        clear_state(user_id)
        await reply(
            "Выберите группу пользователей для рассылки:",
            keyboard=akb.admin_broadcast_groups(),
        )

    elif payload == "adm:bc_all":
        set_state(user_id, "adm_broadcast", bc_group="all")
        count = await db.count_all_user_ids()
        await reply(
            f"Группа: **Все пользователи** ({count} чел.)\n\nОтправьте текст рассылки (можно прикрепить изображение или файл):",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_paid":
        set_state(user_id, "adm_broadcast", bc_group="paid")
        count = await db.count_paid_user_ids()
        await reply(
            f"Группа: **Оплатили тариф** ({count} чел.)\n\nОтправьте текст рассылки (можно прикрепить изображение или файл):",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_no_sub":
        set_state(user_id, "adm_broadcast", bc_group="no_sub")
        count = await db.count_no_sub_user_ids()
        await reply(
            f"Группа: **Без подписки** ({count} чел.)\n\nОтправьте текст рассылки (можно прикрепить изображение или файл):",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_no_paid":
        set_state(user_id, "adm_broadcast", bc_group="no_paid")
        count = await db.count_no_paid_sub_user_ids()
        await reply(
            f"Группа: **Нет платных подписок** ({count} чел.)\n\nОтправьте текст рассылки (можно прикрепить изображение или файл):",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_pending":
        set_state(user_id, "adm_broadcast", bc_group="pending")
        count = await db.count_pending_user_ids()
        await reply(
            f"Группа: **Вызвал оплату, но не оплатил** ({count} чел.)\n\nОтправьте текст рассылки (можно прикрепить изображение или файл):",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_tariff":
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф, подписчикам которого нужно отправить рассылку:",
            keyboard=akb.admin_broadcast_tariff_list(tariffs),
        )

    elif payload.startswith("adm:bc_tariff_pick:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        set_state(user_id, "adm_broadcast", bc_group="tariff", bc_tariff_id=tid)
        count = await db.count_tariff_user_ids(tid)
        await reply(
            f"Группа: **подписчики «{tariff['name']}» ({count} чел.)**\n\nОтправьте текст рассылки (можно прикрепить изображение или файл):",
            keyboard=akb.admin_broadcast_cancel(),
        )



    elif payload == "adm:bc_exclude":
        tariffs = await db.list_tariffs()
        set_state(user_id, "adm_bc_exclude_pick", selected_ids=set(), bc_group="exclude")
        await reply(
            "Выберите тарифы для исключения из рассылки:",
            keyboard=akb.admin_broadcast_exclude_tariff_picker(tariffs, set()),
        )

    elif payload.startswith("adm:bc_exclude_toggle:"):
        tid = int(payload.split(":")[2])
        sd = user_states.get(user_id, {})
        selected = sd.get("selected_ids", set())
        if tid in selected:
            selected.discard(tid)
        else:
            selected.add(tid)
        state_data = user_states.get(user_id, {})
        state_data["selected_ids"] = selected
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тарифы для исключения из рассылки:",
            keyboard=akb.admin_broadcast_exclude_tariff_picker(tariffs, selected),
        )

    elif payload == "adm:bc_exclude_done":
        sd = user_states.get(user_id, {})
        excluded_ids = list(sd.get("selected_ids", []))
        clear_state(user_id)
        set_state(user_id, "adm_broadcast", bc_group="exclude", bc_excluded_ids=excluded_ids)
        count = await db.count_subscribed_excluding_tariffs_user_ids(excluded_ids)
        await reply(
            f"Группа: **Всем кроме выбранных тарифов ({count} чел.)**\n\nОтправьте текст рассылки (можно прикрепить изображение или файл):",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_btn_none":
        await bot.send_message(chat_id, "📤 Рассылка началась")
        sd = user_states.get(user_id, {})
        bc_group = sd.get("bc_group", "all")
        bc_tariff_id = sd.get("bc_tariff_id")
        bc_text = sd.get("bc_text", "")
        bc_media = sd.get("bc_media", [])
        clear_state(user_id)
        if bc_group == "all":
            user_ids = await db.get_all_user_ids()
        elif bc_group == "paid":
            user_ids = await db.get_paid_user_ids()
        elif bc_group == "no_sub":
            user_ids = await db.get_no_sub_user_ids()
        elif bc_group == "no_paid":
            user_ids = await db.get_no_paid_sub_user_ids()
        elif bc_group == "pending":
            user_ids = await db.get_pending_user_ids()
        elif bc_group == "tariff" and bc_tariff_id:
            user_ids = await db.get_tariff_user_ids(bc_tariff_id)
        elif bc_group == "exclude":
            excluded_ids = sd.get("bc_excluded_ids", []) if isinstance(sd, dict) else []
            user_ids = await db.get_subscribed_excluding_tariffs_user_ids(excluded_ids)
        else:
            user_ids = []

        sent = 0
        for uid in user_ids:
            try:
                rendered_text = await _render_broadcast_text(bc_text, uid)
                if bc_media:
                    first = bc_media[0]
                    await bot.forward_attachment(uid, first["type"], first["token"], text=rendered_text)
                    for m in bc_media[1:]:
                        await bot.forward_attachment(uid, m["type"], m["token"])
                else:
                    await bot.send_message(uid, rendered_text)
                sent += 1
            except Exception:
                pass
        await bot.send_message(
            chat_id,
            f"✅ Рассылка завершена.\nОтправлено: **{sent}** из **{len(user_ids)}** пользователей.",
            keyboard=akb.admin_bot_settings(),
        )

    elif payload == "adm:bc_add_btns:yes":
        sd = user_states.get(user_id, {})
        set_state(user_id, "adm_broadcast_wait_input_button",
                  bc_group=sd.get("bc_group"),
                  bc_tariff_id=sd.get("bc_tariff_id"),
                  bc_excluded_ids=sd.get("bc_excluded_ids"),
                  bc_text=sd.get("bc_text"),
                  bc_media=sd.get("bc_media"),
                  bc_buttons=sd.get("bc_buttons", []))
        await reply(
            "📌 Отправьте кнопки в формате:\n"
            "**Текст кнопки - https://ссылка.com**\n\n"
            "Пример:\n"
            "Статья 1 - https://example.com/article-1\n\n"
            "Подойдут `-`, `--` и `—`.\n"
            "Можно отправлять по одной кнопке или весь список сразу (разделяя строками)",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_add_btns:tariff":
        sd = user_states.get(user_id, {})
        buttons = sd.get("bc_buttons", [])
        if len(buttons) >= 5:
            await bot.answer_callback(callback_id, text="Максимум 5 кнопок ⚠️")
            return True
        added_tariff_ids = {
            b["tariff_id"] for b in buttons
            if b.get("kind") == "tariff" and b.get("tariff_id")
        }
        tariffs = await db.list_tariffs()
        await reply(
            "💰 Выберите тариф для кнопки в рассылке:",
            keyboard=akb.admin_broadcast_button_picker(tariffs, added_tariff_ids),
        )

    elif payload == "adm:bc_buttons_menu":
        sd = user_states.get(user_id, {})
        buttons = sd.get("bc_buttons", [])
        if buttons:
            can_add_more = len(buttons) < 5
            await reply(
                _format_broadcast_buttons_message(buttons),
                keyboard=akb.admin_broadcast_button_list(buttons, can_add_more),
            )
        else:
            await reply(
                "➕ Добавить кнопки к рассылке?\n\n"
                "Можно комбинировать тарифы и сторонние ссылки (до 5 кнопок).",
                keyboard=akb.admin_broadcast_add_buttons(),
            )

    elif payload == "adm:bc_add_btns:no":
        await bot.send_message(chat_id, "📤 Рассылка началась")
        sd = user_states.get(user_id, {})
        bc_group = sd.get("bc_group", "all")
        bc_tariff_id = sd.get("bc_tariff_id")
        bc_text = sd.get("bc_text", "")
        bc_media = sd.get("bc_media", [])
        clear_state(user_id)
        if bc_group == "all":
            user_ids = await db.get_all_user_ids()
        elif bc_group == "paid":
            user_ids = await db.get_paid_user_ids()
        elif bc_group == "no_sub":
            user_ids = await db.get_no_sub_user_ids()
        elif bc_group == "no_paid":
            user_ids = await db.get_no_paid_sub_user_ids()
        elif bc_group == "pending":
            user_ids = await db.get_pending_user_ids()
        elif bc_group == "tariff" and bc_tariff_id:
            user_ids = await db.get_tariff_user_ids(bc_tariff_id)
        elif bc_group == "exclude":
            excluded_ids = sd.get("bc_excluded_ids", []) if isinstance(sd, dict) else []
            user_ids = await db.get_subscribed_excluding_tariffs_user_ids(excluded_ids)
        else:
            user_ids = []
        sent = 0
        for uid in user_ids:
            try:
                rendered_text = await _render_broadcast_text(bc_text, uid)
                if bc_media:
                    first = bc_media[0]
                    await bot.forward_attachment(uid, first["type"], first["token"], text=rendered_text)
                    for m in bc_media[1:]:
                        await bot.forward_attachment(uid, m["type"], m["token"])
                else:
                    await bot.send_message(uid, rendered_text)
                sent += 1
            except Exception:
                pass
        await bot.send_message(
            chat_id,
            f"✅ Рассылка завершена.\nОтправлено: **{sent}** из **{len(user_ids)}** пользователей.",
            keyboard=akb.admin_bot_settings(),
        )

    elif payload == "adm:bc_add_btn":
        sd = user_states.get(user_id, {})
        set_state(user_id, "adm_broadcast_wait_input_button",
                  bc_group=sd.get("bc_group"),
                  bc_tariff_id=sd.get("bc_tariff_id"),
                  bc_excluded_ids=sd.get("bc_excluded_ids"),
                  bc_text=sd.get("bc_text"),
                  bc_media=sd.get("bc_media"),
                  bc_buttons=sd.get("bc_buttons", []))
        await reply(
            "📌 Отправьте кнопку в формате:\n"
            "**Текст кнопки - https://ссылка.com**\n"
            "Подойдут `-`, `--` и `—`.",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_add_btn_disabled":
        await bot.answer_callback(callback_id, text="Максимум 5 кнопок достигнут ⚠️")

    elif payload == "adm:bc_send_with_btns":
        await bot.send_message(chat_id, "📤 Рассылка началась")
        sd = user_states.get(user_id, {})
        bc_group = sd.get("bc_group", "all")
        bc_tariff_id = sd.get("bc_tariff_id")
        bc_text = sd.get("bc_text", "")
        bc_media = sd.get("bc_media", [])
        bc_buttons = sd.get("bc_buttons", [])
        
        bc_keyboard = _build_broadcast_keyboard(bc_buttons) if bc_buttons else None
        
        if bc_group == "all":
            user_ids = await db.get_all_user_ids()
        elif bc_group == "paid":
            user_ids = await db.get_paid_user_ids()
        elif bc_group == "no_sub":
            user_ids = await db.get_no_sub_user_ids()
        elif bc_group == "no_paid":
            user_ids = await db.get_no_paid_sub_user_ids()
        elif bc_group == "pending":
            user_ids = await db.get_pending_user_ids()
        elif bc_group == "tariff" and bc_tariff_id:
            user_ids = await db.get_tariff_user_ids(bc_tariff_id)
        elif bc_group == "exclude":
            excluded_ids = sd.get("bc_excluded_ids", []) if isinstance(sd, dict) else []
            user_ids = await db.get_subscribed_excluding_tariffs_user_ids(excluded_ids)
        else:
            user_ids = []
        
        clear_state(user_id)

        sent = 0
        for uid in user_ids:
            try:
                rendered_text = await _render_broadcast_text(bc_text, uid)
                if bc_media:
                    first = bc_media[0]
                    await bot.forward_attachment(uid, first["type"], first["token"], text=rendered_text, keyboard=bc_keyboard)
                    for m in bc_media[1:]:
                        await bot.forward_attachment(uid, m["type"], m["token"])
                else:
                    await bot.send_message(uid, rendered_text, keyboard=bc_keyboard)
                sent += 1
            except Exception:
                pass
        
        await bot.send_message(
            chat_id,
            f"✅ Рассылка завершена.\nОтправлено: **{sent}** из **{len(user_ids)}** пользователей.",
            keyboard=akb.admin_bot_settings(),
        )

    elif payload.startswith("adm:bc_btn_tariff:"):
        tid = int(payload.split(":")[2])
        sd = user_states.get(user_id, {})
        tariff = await db.get_tariff(tid)
        if not tariff:
            await bot.answer_callback(callback_id, text="Тариф не найден")
            return True

        buttons = list(sd.get("bc_buttons", []))
        if len(buttons) >= 5:
            await bot.answer_callback(callback_id, text="Максимум 5 кнопок ⚠️")
            return True
        if any(b.get("kind") == "tariff" and b.get("tariff_id") == tid for b in buttons):
            await bot.answer_callback(callback_id, text="Этот тариф уже добавлен")
            return True

        buttons.append({"kind": "tariff", "text": tariff["name"], "tariff_id": tid})
        can_add_more = len(buttons) < 5

        set_state(user_id, "adm_broadcast_buttons_added",
                  bc_group=sd.get("bc_group"),
                  bc_tariff_id=sd.get("bc_tariff_id"),
                  bc_excluded_ids=sd.get("bc_excluded_ids"),
                  bc_text=sd.get("bc_text"),
                  bc_media=sd.get("bc_media"),
                  bc_buttons=buttons)

        await reply(
            _format_broadcast_buttons_message(buttons),
            keyboard=akb.admin_broadcast_button_list(buttons, can_add_more),
        )

    elif payload == "adm:collect_contacts":
        user_ids = await db.get_all_user_ids()
        sent = 0
        for uid in user_ids:
            try:
                await bot.send_message(
                    uid,
                    "📱 Пожалуйста, поделитесь вашим номером телефона, "
                    "нажав кнопку ниже 👇",
                    keyboard=_contact_request_kb(),
                )
                sent += 1
            except Exception:
                pass
        await reply(
            f"✅ Запрос на получение номера телефона отправлен {sent} пользователям.",
            keyboard=akb.admin_bot_settings(),
        )

    elif payload.startswith("adm:reply_feedback:"):
        target = int(payload.split(":")[2])
        set_state(user_id, "adm_reply_feedback", target_user_id=target)
        await bot.send_message(
            chat_id,
            "Ответьте на вопрос. Это может быть текст, фото, видео "
            "или любое другое медиа вложение:",
            keyboard=akb.admin_cancel_feedback_reply(),
        )

    elif payload == "adm:cancel_feedback_reply":
        clear_state(user_id)
        await bot.send_message(chat_id, "Ответ отменён.")

    elif payload.startswith("adm:ban_user:"):
        target = int(payload.split(":")[2])
        await db.ban_user(target)
        target_user = await db.get_user(target)
        name = ""
        if target_user:
            name = f"{target_user['first_name']} {target_user['last_name']}".strip()
        await bot.send_message(
            chat_id,
            f"🚫 Пользователь {user_link(name, target)} заблокирован.",
            fmt="markdown",
        )


    elif payload == "adm:payment_methods":
        clear_state(user_id)
        methods = await db.list_payment_methods()
        await reply(
            "💳 Способы оплаты.\n\nНажмите на метод для управления или добавьте новый.",
            keyboard=akb.admin_payment_methods_list(methods),
        )

    elif payload == "adm:add_pay_method":
        import payments as pay_mod
        providers = pay_mod.provider_names()
        await reply(
            "Выберите платёжную систему:",
            keyboard=akb.admin_payment_provider_list(providers),
        )

    elif payload.startswith("adm:pay_provider:"):
        provider_key = payload.split(":", 2)[2]
        set_state(user_id, "adm_pay_name", provider=provider_key)
        await reply(
            "Введите название способа оплаты\n(как он будет отображаться пользователям):",
            keyboard=akb.admin_payment_cancel(),
        )

    elif payload.startswith("adm:pay_detail:"):
        mid = int(payload.split(":")[2])
        method = await db.get_payment_method(mid)
        if not method:
            await reply("Метод не найден.", keyboard=akb.admin_payment_methods_list(await db.list_payment_methods()))
            return
        status = "✅ Включён" if method["is_active"] else "❌ Выключен"
        text = (
            f"💳 **{method['name']}**\n"
            f"Провайдер: {method['provider']}\n"
            f"Shop ID: {method['shop_id']}\n"
            f"Статус: {status}"
        )
        await reply(text, keyboard=akb.admin_payment_detail(mid, method["is_active"]))

    elif payload.startswith("adm:toggle_pay:"):
        mid = int(payload.split(":")[2])
        method = await db.toggle_payment_method(mid)
        if method:
            status = "включён ✅" if method["is_active"] else "выключен ❌"
            await reply(f"Метод «{method['name']}» — {status}",
                        keyboard=akb.admin_payment_detail(mid, method["is_active"]))

    elif payload.startswith("adm:del_pay:"):
        mid = int(payload.split(":")[2])
        method = await db.get_payment_method(mid)
        await reply(
            f"Удалить способ оплаты «{method['name']}»?",
            keyboard=akb.admin_confirm_pay_delete(mid),
        )

    elif payload.startswith("adm:del_pay_confirm:"):
        mid = int(payload.split(":")[2])
        await db.delete_payment_method(mid)
        methods = await db.list_payment_methods()
        await reply("✅ Способ оплаты удалён.", keyboard=akb.admin_payment_methods_list(methods))


    elif payload == "adm:editing_menu":
        clear_state(user_id)
        await reply(
            "Выберите что вы хотите отредактировать:",
            keyboard=akb.admin_editing_menu(),
        )

    elif payload == "adm:button_texts":
        clear_state(user_id)
        await reply(
            "🔤 Редактирование текста кнопок Личного кабинета.\n"
            "Выберите кнопку для изменения:",
            keyboard=akb.admin_button_texts_list(db.BUTTON_TEXT_LABELS),
        )

    elif payload.startswith("adm:edit_btn:"):
        key = payload.split(":", 2)[2]
        label = db.BUTTON_TEXT_LABELS.get(key, key)
        current = await db.get_bot_text(key)
        set_state(user_id, "adm_edit_btn_text", text_key=key)
        await reply(
            f"**{label}**\n\nТекущий текст кнопки:\n{current}\n\n"
            "Отправьте новый текст:",
            keyboard=akb.admin_edit_btn_back(),
        )

    elif payload == "adm:desc_texts":
        clear_state(user_id)
        await reply(
            "📝 Редактирование описаний.\n"
            "Выберите описание для изменения:",
            keyboard=akb.admin_desc_texts_list(db.DESC_TEXT_LABELS),
        )

    elif payload.startswith("adm:edit_desc:"):
        key = payload.split(":", 2)[2]
        label = db.DESC_TEXT_LABELS.get(key, key)
        current = await db.get_bot_text(key)
        preview = current[:300] + "…" if len(current) > 300 else current
        set_state(user_id, "adm_edit_desc_text", text_key=key)
        await reply(
            f"**{label}**\n\nТекущий текст:\n{preview}\n\n"
            "Отправьте новый текст:",
            keyboard=akb.admin_edit_desc_back(),
        )


    elif payload == "adm:bot_texts":
        clear_state(user_id)
        await reply(
            "Редактирование текстов бота.\nВыберите текст для изменения:",
            keyboard=akb.admin_bot_texts_list(db.BOT_TEXT_LABELS),
        )

    elif payload.startswith("adm:edit_text:"):
        key = payload.split(":", 2)[2]
        label = db.BOT_TEXT_LABELS.get(key, key)
        current = await db.get_bot_text(key)
        preview = current[:300] + "…" if len(current) > 300 else current
        set_state(user_id, "adm_edit_bot_text", text_key=key)
        await reply(
            f"**{label}**\n\nТекущий текст:\n{preview}\n\n"
            "Отправьте новый текст:",
            keyboard=akb.admin_bot_text_back(),
        )


    elif payload == "adm:promo_menu":
        clear_state(user_id)
        await reply(
            "**Настройки промокодов**\n\n"
            "Выберите нужный раздел промокодов ниже или создайте новый:",
            keyboard=akb.admin_promo_menu(),
        )

    elif payload == "adm:promo_general":
        promos = await db.list_promos("general")
        if promos:
            await reply("Общие промокоды:", keyboard=akb.admin_promo_list(promos))
        else:
            set_state(user_id, "adm_promo_name", promo_type="general")
            await reply(
                "Отправьте боту новый промокод\n(например: discount30, PROMO20)",
                keyboard=akb.admin_promo_back(),
            )

    elif payload == "adm:promo_broadcast":
        promos = await db.list_promos("broadcast")
        if promos:
            await reply("Промокоды из рассылок:", keyboard=akb.admin_promo_list(promos))
        else:
            await reply("Промокодов из рассылок пока нет.", keyboard=akb.admin_promo_back())

    elif payload == "adm:promo_activation":
        promos = await db.list_promos("activation")
        if promos:
            await reply("Промокоды активационных ссылок:", keyboard=akb.admin_promo_list(promos))
        else:
            await reply("Промокодов активационных ссылок пока нет.", keyboard=akb.admin_promo_back())

    elif payload == "adm:promo_create":
        set_state(user_id, "adm_promo_name", promo_type="general")
        await reply(
            "Отправьте боту новый промокод\n(например: discount30, PROMO20)",
            keyboard=akb.admin_promo_back(),
        )

    elif payload == "adm:promo_create_group":
        set_state(user_id, "adm_promo_group_name")
        await reply(
            "Введите название группы промокодов:",
            keyboard=akb.admin_promo_back(),
        )

    elif payload.startswith("adm:promo_open:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        if not promo:
            await reply("Промокод не найден.")
            return True
        await _show_promo_detail(reply, promo)

    elif payload.startswith("adm:promo_tariffs:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        tariffs = await db.list_tariffs()
        selected = set()
        if promo.get("allowed_tariffs"):
            selected = {int(x) for x in promo["allowed_tariffs"].split(",") if x.strip()}
        set_state(user_id, "adm_promo_tariff_pick", promo_id=pid, selected_tariffs=selected)
        await reply(
            f"Промокод **{promo['code']}**\nВыберите тарифы, к которым можно применить промокод:",
            keyboard=akb.admin_promo_tariff_picker(tariffs, selected, pid),
        )

    elif payload.startswith("adm:promo_toggle_t:"):
        parts = payload.split(":")
        pid = int(parts[2])
        tid = int(parts[3])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_tariffs", set())
        if tid in selected:
            selected.discard(tid)
        else:
            selected.add(tid)
        state_data["selected_tariffs"] = selected
        tariffs = await db.list_tariffs()
        promo = await db.get_promo(pid)
        await reply(
            f"Промокод **{promo['code']}**\nВыберите тарифы:",
            keyboard=akb.admin_promo_tariff_picker(tariffs, selected, pid),
        )

    elif payload.startswith("adm:promo_save_t:"):
        pid = int(payload.split(":")[2])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_tariffs", set())
        val = ",".join(str(x) for x in selected) if selected else None
        await db.update_promo(pid, allowed_tariffs=val)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await reply("Разрешённые тарифы обновлены ✅")
        await _show_promo_detail(reply, promo)

    elif payload.startswith("adm:promo_edit_max:"):
        pid = int(payload.split(":")[2])
        set_state(user_id, "adm_promo_edit_max", promo_id=pid)
        await reply(
            "Введите максимальное кол-во активаций (0 = безлимит):",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_edit_per_user:"):
        pid = int(payload.split(":")[2])
        set_state(user_id, "adm_promo_edit_per_user", promo_id=pid)
        await reply(
            "Введите максимальное кол-во активаций на одного человека:",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_edit_group:"):
        pid = int(payload.split(":")[2])
        set_state(user_id, "adm_promo_edit_group", promo_id=pid)
        await reply(
            "Введите название группы разрешённых пользователей\n(или /empty чтобы сбросить):",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_edit_expiry:"):
        pid = int(payload.split(":")[2])
        set_state(user_id, "adm_promo_edit_expiry", promo_id=pid)
        await reply(
            "Введите дату окончания в формате ДД.ММ.ГГГГ ЧЧ:ММ\n(или /empty для безлимитного срока):",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_allowed_users:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        users_str = promo.get("allowed_users") or "Не ограничено (все пользователи)"
        await reply(
            f"Промокод **{promo['code']}**\n\n"
            f"Разрешённые пользователи:\n{users_str}",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_activations:"):
        pid = int(payload.split(":")[2])
        await _send_promo_activations_xlsx(bot, chat_id, pid)
        promo = await db.get_promo(pid)
        await bot.send_message(
            chat_id, f"👆 Активации промокода **{promo['code']}**",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_delete:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        await reply(
            f"⚠️ Удалить промокод **{promo['code']}**?",
            keyboard=akb.admin_promo_confirm_delete(pid),
        )

    elif payload.startswith("adm:promo_confirm_del:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        code = promo["code"] if promo else "?"
        await db.delete_promo(pid)
        await reply(
            f"Промокод **{code}** удалён 🗑",
            keyboard=akb.admin_promo_menu(),
        )

    else:
        await reply("Неизвестная команда админки")

    return True


async def handle_admin_message(
    bot: MaxBot,
    user_id: int,
    chat_id: int,
    text: str,
    attachments: list | None = None,
) -> bool:
    """Обрабатывает текстовые сообщения в контексте админ-FSM.
    Возвращает True если обработал.
    """
    state = get_state(user_id)
    if not state.startswith("adm_"):
        return False

    state_data = user_states.get(user_id, {})

    if state == "adm_add_resource":
        try:
            resource_chat_id = int(text.strip())
        except ValueError:
            await bot.send_message(
                chat_id,
                "Введите числовой `chat_id` чата или канала.",
                keyboard=akb.admin_back_to_settings_menu(),
            )
            return True

        try:
            info = await bot.get_chat_info(resource_chat_id)
        except Exception as e:
            await bot.send_message(
                chat_id,
                f"❌ Не удалось получить ресурс: {e}",
                keyboard=akb.admin_back_to_settings_menu(),
            )
            return True

        status = info.get("status")
        if info.get("chat_id") and int(info["chat_id"]) != resource_chat_id:
            resource_chat_id = int(info["chat_id"])
        if not info.get("chat_id") and not info.get("title"):
            await bot.send_message(
                chat_id,
                f"❌ MAX не вернул информацию о ресурсе. Ответ: {info}",
                keyboard=akb.admin_back_to_settings_menu(),
            )
            return True
        if status and status != "active":
            await bot.send_message(
                chat_id,
                f"❌ Ресурс найден, но бот не активен в нём (status={status}).",
                keyboard=akb.admin_back_to_settings_menu(),
            )
            return True

        title = info.get("title") or str(resource_chat_id)
        await db.upsert_bot_chat(
            resource_chat_id,
            title=title,
            link=info.get("link", ""),
            is_channel=info.get("type") == "channel" or info.get("is_channel", False),
        )

        clear_state(user_id)
        chats = await db.get_all_bot_chats()
        usage = await db.get_resource_usage()
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        for cid, _ in usage.items():
            if cid not in chat_ids_in_list:
                chats.append({"chat_id": cid, "title": f"❓ {cid} (недоступен)"})
        await bot.send_message(
            chat_id,
            f"✅ Ресурс «{title}» добавлен.\n\n📋 Ресурсы бота (чаты/каналы).",
            keyboard=akb.admin_resources_list(chats, usage),
        )
        return True

    if state == "adm_create_name":
        set_state(user_id, "adm_create_price", tariff_name=text)
        await bot.send_message(
            chat_id,
            f"Название тарифа — {text}\n \nВведите цену или сделайте тариф бесплатным",
            keyboard=akb.admin_create_price(),
        )
        return True

    if state == "adm_create_price":
        try:
            price = float(text.replace(",", ".").replace(" ", ""))
        except ValueError:
            await bot.send_message(chat_id, "Введите число (цену в рублях):", keyboard=akb.admin_create_price())
            return True
        name = state_data.get("tariff_name", "")
        set_state(user_id, "adm_create_resources", tariff_name=name, tariff_price=price, is_free=False)
        await bot.send_message(
            chat_id,
            f"Название тарифа — {name}\n \nЦена: {price}₽\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )
        return True

    if state == "adm_create_duration_custom":
        name = state_data.get("tariff_name", "")
        price = state_data.get("tariff_price", 0)
        is_free = state_data.get("is_free", False)

        dur_text = text.strip()
        duration_minutes = _parse_duration_to_minutes(dur_text)
        set_state(user_id, "adm_create_resources",
                  tariff_name=name, tariff_price=price, is_free=is_free,
                  duration_days=None, duration_minutes=duration_minutes,
                  duration_text=dur_text)
        price_str = "бесплатно" if is_free else f"{price}₽"
        await bot.send_message(
            chat_id,
            f"Название тарифа — {name}\n \n"
            f"Цена: {price_str}\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, "
            "нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )
        return True

    if state == "adm_create_duration":
        try:
            days = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите число дней:", keyboard=akb.admin_create_duration())
            return True
        name = state_data.get("tariff_name", "")
        price = state_data.get("tariff_price", 0)
        is_free = state_data.get("is_free", False)
        dur_text = f"{days} дн."
        set_state(user_id, "adm_create_resources",
                  tariff_name=name, tariff_price=price, is_free=is_free,
                  duration_days=days, duration_text=dur_text)
        price_str = "бесплатно" if is_free else f"{price}₽"
        await bot.send_message(
            chat_id,
            f"Название тарифа — {name}\n \n"
            f"Цена: {price_str}\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, "
            "нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )
        return True

    if state == "adm_create_category":
        await db.create_category(text.strip())
        clear_state(user_id)
        tariffs = await db.list_tariffs()
        await bot.send_message(
            chat_id,
            f"Категория «{text.strip()}» создана ✅\n \nСписок ваших категорий и тарифов",
            keyboard=akb.admin_tariff_list(tariffs),
        )
        return True

    if state == "adm_gift_wait_file":
        atts = attachments or []
        file_token = ""
        file_name = ""
        for att in atts:
            if not isinstance(att, dict):
                continue
            if att.get("type") != "file":
                continue
            payload = att.get("payload") or {}
            file_token = payload.get("token") or payload.get("file_token") or ""
            file_name = payload.get("name") or payload.get("file_name") or ""
            if file_token:
                break

        if not file_token:
            await bot.send_message(chat_id, "Пришлите именно файл (вложение).")
            return True

        selected: set[int] = state_data.get("selected_tariffs", set())
        gift = await db.create_gift_file(file_token=file_token, file_name=file_name, tariff_ids=list(selected))
        clear_state(user_id)
        if gift:
            await bot.send_message(chat_id, "✅ Гифт файл сохранён.", keyboard=akb.admin_main())
        else:
            await bot.send_message(chat_id, "❌ Не удалось сохранить гифт файл.", keyboard=akb.admin_main())
        return True

    if state == "adm_tariff_gift_wait_file":
        atts = attachments or []
        file_token = ""
        file_name = ""
        for att in atts:
            if not isinstance(att, dict):
                continue
            if att.get("type") != "file":
                continue
            att_payload = att.get("payload") or {}
            file_token = att_payload.get("token") or att_payload.get("file_token") or ""
            file_name = att_payload.get("name") or att_payload.get("file_name") or ""
            if file_token:
                break

        if not file_token:
            await bot.send_message(chat_id, "Пришлите именно файл (вложение).")
            return True

        tid = state_data.get("tariff_id")
        gift = await db.create_gift_file(file_token=file_token, file_name=file_name, tariff_ids=[tid])
        clear_state(user_id)
        if gift:
            tariff = await db.get_tariff(tid)
            await bot.send_message(chat_id, "✅ Бонусный файл добавлен.", keyboard=akb.admin_tariff_gifts_menu(tid))
        else:
            await bot.send_message(chat_id, "❌ Не удалось сохранить файл.")
        return True

    if state == "adm_edit_name":
        tid = state_data.get("tariff_id")
        await db.update_tariff(tid, name=text.strip())
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, f"Название обновлено ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_desc":
        tid = state_data.get("tariff_id")
        desc = "" if text.strip().lower() == "delete" else text.strip()
        await db.update_tariff(tid, description=desc)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Описание обновлено ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_price_std":
        tid = state_data.get("tariff_id")
        try:
            price = float(text.replace(",", ".").replace(" ", ""))
        except ValueError:
            await bot.send_message(chat_id, "Введите число:")
            return True
        await db.update_tariff(tid, price=price, is_free=(price == 0))
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Цена обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_price_old":
        tid = state_data.get("tariff_id")
        val = _parse_nullable_float(text)
        await db.update_tariff(tid, old_price=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Старая цена обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_price_renew":
        tid = state_data.get("tariff_id")
        val = _parse_nullable_float(text)
        await db.update_tariff(tid, renewal_price=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Цена продления обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_price_active":
        tid = state_data.get("tariff_id")
        val = _parse_nullable_float(text)
        await db.update_tariff(tid, active_renewal_price=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Цена продления активной подписки обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_duration":
        tid = state_data.get("tariff_id")
        try:
            days = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите число дней:")
            return True
        dur_text = f"{days} дн."
        await db.update_tariff(tid, duration_days=days, duration_text=dur_text, duration_minutes=None, end_date=None)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Продолжительность обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_duration_custom":
        tid = state_data.get("tariff_id")
        dur_text = text.strip()
        duration_minutes = _parse_duration_to_minutes(dur_text)
        if not duration_minutes:
            await bot.send_message(chat_id, "Не удалось распознать. Введите например: 48ч или 120м")
            return True
        await db.update_tariff(
            tid,
            duration_days=None,
            duration_minutes=duration_minutes,
            duration_text=dur_text,
            end_date=None,
        )
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Продолжительность обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_start_day":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, start_day=None)
        else:
            try:
                day = int(text.strip())
            except ValueError:
                dt = _parse_datetime(text.strip())
                if not dt:
                    await bot.send_message(chat_id, "Введите число (день) или дату (16.03.2026 07:54):")
                    return True
                day = dt.day
            await db.update_tariff(tid, start_day=day)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "День начала обновлён ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_start_date":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, start_date=None)
        else:
            dt = _parse_datetime(text.strip())
            if not dt:
                await bot.send_message(chat_id, "Формат: 16.03.2026 07:54")
                return True
            await db.update_tariff(tid, start_date=dt)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Дата начала обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_end_date":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, end_date=None)
        else:
            dt = _parse_datetime(text.strip())
            if not dt:
                await bot.send_message(chat_id, "Формат: 16.03.2026 07:54")
                return True
            await db.update_tariff(
                tid,
                end_date=dt,
                duration_days=None,
                duration_minutes=None,
                duration_text="",
            )
            await db.update_active_purchases_expiry(tid, dt)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Дата конца обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_check_name":
        tid = state_data.get("tariff_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_tariff(tid, check_name=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Название в чеке обновлено ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_reject":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, rejection_interval=None)
        else:
            try:
                val = int(text.strip())
            except ValueError:
                await bot.send_message(chat_id, "Введите число минут:")
                return True
            await db.update_tariff(tid, rejection_interval=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Интервал отклонений обновлён ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_success":
        tid = state_data.get("tariff_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_tariff(tid, success_text=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Текст успешной покупки обновлён ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_limit":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, activation_limit=None)
        else:
            try:
                val = int(text.strip())
            except ValueError:
                await bot.send_message(chat_id, "Введите число:")
                return True
            await db.update_tariff(tid, activation_limit=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Лимит активаций обновлён ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_media":
        tid = state_data.get("tariff_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_tariff(tid, media_url=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Медиа обновлено ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_warmup_add_text":
        tid = state_data.get("tariff_id")
        set_state(user_id, "adm_warmup_add_delay", tariff_id=tid, warmup_text=text.strip())
        await bot.send_message(
            chat_id,
            "Введите время отправки после начала оплаты.\nФормат: например 2ч, 30м, 48ч",
            keyboard=akb.admin_warmup_cancel(tid),
        )
        return True

    if state == "adm_warmup_add_delay":
        tid = state_data.get("tariff_id")
        delay_minutes = _parse_duration_to_minutes(text.strip())
        if not delay_minutes:
            await bot.send_message(chat_id, "Не удалось распознать. Введите например: 2ч или 30м")
            return True
        warmup_text = state_data.get("warmup_text", "")
        set_state(user_id, "adm_warmup_add_buttons",
                  tariff_id=tid,
                  warmup_text=warmup_text,
                  warmup_delay_minutes=delay_minutes,
                  warmup_buttons=[])
        await bot.send_message(
            chat_id,
            "➕ Добавить кнопки к догревающей рассылке?\n\n"
            "Можно комбинировать тарифы и сторонние ссылки (до 5 кнопок).\n"
            "• **Сторонняя ссылка** — переход на URL\n"
            "• **Кнопка тарифа** — оформление тарифа в боте\n"
            "• **Без кнопки** — только текст/медиа",
            keyboard=akb.admin_warmup_add_buttons(tid),
        )
        return True

    if state == "adm_warmup_edit_text":
        mid = state_data.get("message_id")
        tid = state_data.get("tariff_id")
        await db.update_warmup_message(mid, text=text.strip())
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        messages = await db.get_warmup_messages(tid)
        await bot.send_message(chat_id, "Текст обновлён ✅")
        await bot.send_message(
            chat_id,
            f"🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )
        return True

    if state == "adm_warmup_edit_media":
        mid = state_data.get("message_id")
        tid = state_data.get("tariff_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_warmup_message(mid, media_url=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        messages = await db.get_warmup_messages(tid)
        await bot.send_message(chat_id, "Медиа обновлено ✅")
        await bot.send_message(
            chat_id,
            f"🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )
        return True

    if state == "adm_warmup_edit_delay":
        mid = state_data.get("message_id")
        tid = state_data.get("tariff_id")
        delay_minutes = _parse_duration_to_minutes(text.strip())
        if not delay_minutes:
            await bot.send_message(chat_id, "Не удалось распознать. Введите например: 2ч или 30м")
            return True
        await db.update_warmup_message(mid, delay_minutes=delay_minutes)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        messages = await db.get_warmup_messages(tid)
        await bot.send_message(chat_id, "Время отправки обновлено ✅")
        await bot.send_message(
            chat_id,
            f"🔥 Догревающие рассылки тарифа «{tariff['name']}»",
            keyboard=akb.admin_warmup_list(tid, messages, tariff.get("warmup_order_mode") or "sequential"),
        )
        return True

    if state in ("adm_warmup_wait_input_button", "adm_warmup_add_wait_input_button", "adm_warmup_add_buttons_added"):
        parsed_buttons, invalid_line = _parse_broadcast_button_lines(text)
        if invalid_line is not None or not parsed_buttons:
            await bot.send_message(
                chat_id,
                "❌ Неверный формат! Используйте:\n"
                "**Текст кнопки - URL**\n\n"
                "Пример: Купить курс - https://example.com/buy\n"
                "Можно отправлять по одной кнопке или списком, по одной строке на кнопку.",
                keyboard=akb.admin_warmup_cancel(state_data.get("tariff_id")),
            )
            return True

        buttons = list(state_data.get("warmup_buttons", []))
        if len(buttons) + len(parsed_buttons) > 5:
            await bot.send_message(
                chat_id,
                "❌ Максимум 5 кнопок в рассылке.",
                keyboard=akb.admin_warmup_cancel(state_data.get("tariff_id")),
            )
            return True

        buttons.extend(parsed_buttons)
        tid = state_data.get("tariff_id")
        mid = state_data.get("message_id")
        if state == "adm_warmup_wait_input_button" and mid:
            await db.update_warmup_message(mid, buttons=buttons)
            clear_state(user_id)
            await bot.send_message(
                chat_id,
                format_inline_buttons_message(buttons),
                keyboard=akb.admin_warmup_button_list(mid, tid, buttons, len(buttons) < 5),
            )
            return True

        set_state(user_id, "adm_warmup_add_buttons_added",
                  tariff_id=tid,
                  warmup_text=state_data.get("warmup_text"),
                  warmup_delay_minutes=state_data.get("warmup_delay_minutes"),
                  warmup_buttons=buttons)
        await bot.send_message(
            chat_id,
            format_inline_buttons_message(buttons),
            keyboard=akb.admin_warmup_button_list(None, tid, buttons, len(buttons) < 5),
        )
        return True

    if state == "adm_profile_input":
        try:
            target_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой id пользователя:")
            return True
        profile = await db.user_profile(target_id)
        if not profile:
            await bot.send_message(
                chat_id, "Пользователь не найден. Убедитесь что он запускал бота.",
                keyboard=akb.admin_back_subscribers(),
            )
            clear_state(user_id)
            return True
        clear_state(user_id)
        await _send_user_profile(bot, chat_id, profile)
        return True

    if state == "adm_search_id_input":
        try:
            target_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой ID пользователя:")
            return True
        profile = await db.user_profile(target_id)
        if not profile:
            await bot.send_message(
                chat_id, "Пользователь не найден. Убедитесь что он запускал бота.",
                keyboard=akb.admin_back_subscribers(),
            )
            clear_state(user_id)
            return True
        clear_state(user_id)
        await _send_user_profile_with_logs(bot, chat_id, profile)
        return True

    if state == "adm_grant_user":
        try:
            target_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой user\\_ID:")
            return True
        tid = state_data.get("tariff_id")
        target = await db.get_user(target_id)
        if not target:
            await bot.send_message(
                chat_id,
                "Пользователь не найден. Он должен сначала написать /start боту.",
                keyboard=akb.admin_back_subscribers(),
            )
            clear_state(user_id)
            return True
        tariff = await db.get_tariff(tid)
        await db.grant_subscription(target_id, tid)
        clear_state(user_id)
        full_name = f"{target['first_name']} {target['last_name']}".strip()
        await bot.send_message(
            chat_id,
            f"✅ Подписка «{tariff['name']}» выдана пользователю "
            f"{user_link(full_name, target_id)}",
            keyboard=akb.admin_subscribers(),
            fmt="markdown",
        )
        return True

    if state == "adm_revoke_user":
        try:
            target_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой user\\_ID:")
            return True
        tid = state_data.get("tariff_id")
        tariff = await db.get_tariff(tid)
        await db.revoke_subscription(target_id, tid)
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"✅ Подписка «{tariff['name']}» обнулена у пользователя id={target_id}",
            keyboard=akb.admin_subscribers(),
        )
        return True

    if state == "adm_transfer_from":
        try:
            from_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой user\\_ID:")
            return True
        tid = state_data.get("tariff_id")
        set_state(user_id, "adm_transfer_to", tariff_id=tid, from_user_id=from_id)
        await bot.send_message(
            chat_id,
            f"Теперь отправьте user\\_ID пользователя, **которому** передать подписку:",
            keyboard=akb.admin_back_subscribers(),
        )
        return True

    if state == "adm_transfer_to":
        try:
            to_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой user\\_ID:")
            return True
        tid = state_data.get("tariff_id")
        from_id = state_data.get("from_user_id")
        tariff = await db.get_tariff(tid)
        ok = await db.transfer_subscription(from_id, to_id, tid)
        clear_state(user_id)
        if ok:
            await bot.send_message(
                chat_id,
                f"✅ Подписка «{tariff['name']}» передана от id={from_id} к id={to_id}",
                keyboard=akb.admin_subscribers(),
            )
        else:
            await bot.send_message(
                chat_id,
                f"❌ У пользователя id={from_id} нет активной подписки «{tariff['name']}»",
                keyboard=akb.admin_subscribers(),
            )
        return True

    if state == "adm_reply_feedback":
        target_id = state_data.get("target_user_id")
        media_atts = [
            att for att in (attachments or [])
            if att.get("type") in ("image", "file", "video", "audio")
               and att.get("payload", {}).get("token")
        ]
        if text:
            await bot.send_message(
                target_id,
                await db.get_bot_text("feedback_reply", user_id=target_id, reply=text),
            )
        for att in media_atts:
            att_type = att.get("type", "file")
            token = att.get("payload", {}).get("token", "")
            if token:
                await bot.forward_attachment(target_id, att_type, token)
        set_state(target_id, "waiting_feedback")
        clear_state(user_id)
        await bot.send_message(
            chat_id, "✅ Ответ отправлен пользователю.",
            keyboard=akb.admin_main(),
        )
        return True

    if state == "adm_msg_user":
        target_id = state_data.get("target_user_id")
        await bot.send_message(target_id, text)
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"✅ Сообщение отправлено пользователю id={target_id}",
            keyboard=akb.admin_subscribers(),
        )
        return True


    if state == "adm_broadcast":
        bc_group = state_data.get("bc_group", "all")
        bc_tariff_id = state_data.get("bc_tariff_id")
        bc_excluded_ids = state_data.get("bc_excluded_ids", [])
        bc_media = [
            {"type": att.get("type"), "token": att.get("payload", {}).get("token")}
            for att in (attachments or [])
            if att.get("type") in ("image", "file", "video", "audio")
            and att.get("payload", {}).get("token")
        ]
        set_state(user_id, "adm_broadcast_add_buttons",
                  bc_group=bc_group, bc_tariff_id=bc_tariff_id, bc_excluded_ids=bc_excluded_ids, bc_text=text, bc_media=bc_media, bc_buttons=[])
        await bot.send_message(
            chat_id,
            "➕ Добавить кнопки к рассылке?\n\n"
            "Можно комбинировать тарифы и сторонние ссылки (до 5 кнопок).\n"
            "• **Сторонняя ссылка** — переход на URL\n"
            "• **Кнопка тарифа** — оформление тарифа в боте\n"
            "• **Без кнопки** — только текст/медиа",
            keyboard=akb.admin_broadcast_add_buttons(),
        )
        return True

    if state in ("adm_broadcast_wait_input_button", "adm_broadcast_buttons_added"):
        parsed_buttons, invalid_line = _parse_broadcast_button_lines(text)
        if invalid_line is not None or not parsed_buttons:
            await bot.send_message(
                chat_id,
                "❌ Неверный формат! Используйте:\n"
                "**Текст кнопки - URL**\n\n"
                "Пример: Купить курс - https://example.com/buy\n"
                "Можно отправлять по одной кнопке или списком, по одной строке на кнопку.",
                keyboard=akb.admin_broadcast_cancel(),
            )
            return True

        buttons = list(state_data.get("bc_buttons", []))
        if len(buttons) + len(parsed_buttons) > 5:
            await bot.send_message(
                chat_id,
                "❌ Максимум 5 кнопок в рассылке.",
                keyboard=akb.admin_broadcast_cancel(),
            )
            return True

        buttons.extend(parsed_buttons)
        can_add_more = len(buttons) < 5

        set_state(user_id, "adm_broadcast_buttons_added",
                  bc_group=state_data.get("bc_group"),
                  bc_tariff_id=state_data.get("bc_tariff_id"),
                  bc_excluded_ids=state_data.get("bc_excluded_ids"),
                  bc_text=state_data.get("bc_text"),
                  bc_media=state_data.get("bc_media"),
                  bc_buttons=buttons)

        await bot.send_message(
            chat_id,
            _format_broadcast_buttons_message(buttons),
            keyboard=akb.admin_broadcast_button_list(buttons, can_add_more),
        )
        return True


    if state == "adm_edit_bot_text":
        key = state_data.get("text_key")
        await db.set_bot_text(key, text.strip())
        clear_state(user_id)
        label = db.BOT_TEXT_LABELS.get(key, key)
        await bot.send_message(
            chat_id,
            f"✅ Текст «{label}» обновлён.",
            keyboard=akb.admin_bot_texts_list(db.BOT_TEXT_LABELS),
        )
        return True

    if state == "adm_edit_btn_text":
        key = state_data.get("text_key")
        await db.set_bot_text(key, text.strip())
        clear_state(user_id)
        label = db.BUTTON_TEXT_LABELS.get(key, key)
        await bot.send_message(
            chat_id,
            f"✅ Текст кнопки «{label}» обновлён на: {text.strip()}",
            keyboard=akb.admin_button_texts_list(db.BUTTON_TEXT_LABELS),
        )
        return True

    if state == "adm_edit_desc_text":
        key = state_data.get("text_key")
        await db.set_bot_text(key, text.strip())
        clear_state(user_id)
        label = db.DESC_TEXT_LABELS.get(key, key)
        await bot.send_message(
            chat_id,
            f"✅ Описание «{label}» обновлено.",
            keyboard=akb.admin_desc_texts_list(db.DESC_TEXT_LABELS),
        )
        return True


    if state == "adm_promo_name":
        code = text.strip()
        existing = await db.get_promo_by_code(code)
        if existing:
            await bot.send_message(
                chat_id, f"❌ Промокод «{code}» уже существует. Введите другой:",
                keyboard=akb.admin_promo_back(),
            )
            return True
        promo_type = state_data.get("promo_type", "general")
        set_state(user_id, "adm_promo_discount", promo_code=code, promo_type=promo_type)
        await bot.send_message(
            chat_id,
            f"Промокод — **{code}**\n\n"
            "Отправьте боту процент скидки на покупку тарифов (1-100).",
            keyboard=akb.admin_promo_back(),
        )
        return True

    if state == "adm_promo_discount":
        try:
            pct = int(text.strip())
            if not 1 <= pct <= 100:
                raise ValueError
        except ValueError:
            await bot.send_message(chat_id, "Введите число от 1 до 100:")
            return True
        code = state_data.get("promo_code", "")
        promo_type = state_data.get("promo_type", "general")
        set_state(user_id, "adm_promo_max_act",
                  promo_code=code, promo_discount=pct, promo_type=promo_type)
        await bot.send_message(
            chat_id,
            f"**Создание промокода**\n\n"
            f"Промокод — **{code}**\n"
            f"Процент скидки — **{pct}%**\n"
            "Отправьте боту нужное количество активаций "
            "либо отправьте 0 для его безлимитного числа.",
            keyboard=akb.admin_promo_back(),
        )
        return True

    if state == "adm_promo_max_act":
        try:
            max_act = int(text.strip())
            if max_act < 0:
                raise ValueError
        except ValueError:
            await bot.send_message(chat_id, "Введите неотрицательное число:")
            return True
        code = state_data.get("promo_code", "")
        pct = state_data.get("promo_discount", 0)
        promo_type = state_data.get("promo_type", "general")
        promo = await db.create_promo(code, pct, max_act, promo_type)
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"Промокод **{code}** создан ✅",
            keyboard=akb.admin_promo_created(promo["id"]),
        )
        return True

    if state == "adm_promo_group_name":
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"✅ Группа промокодов «{text.strip()}» создана.",
            keyboard=akb.admin_promo_menu(),
        )
        return True

    if state == "adm_promo_edit_max":
        try:
            val = int(text.strip())
            if val < 0:
                raise ValueError
        except ValueError:
            await bot.send_message(chat_id, "Введите неотрицательное число:")
            return True
        pid = state_data.get("promo_id")
        await db.update_promo(pid, max_activations=val)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await bot.send_message(chat_id, "✅ Кол-во активаций обновлено.")
        await _send_promo_detail(bot, chat_id, promo)
        return True

    if state == "adm_promo_edit_per_user":
        try:
            val = int(text.strip())
            if val < 1:
                raise ValueError
        except ValueError:
            await bot.send_message(chat_id, "Введите число ≥ 1:")
            return True
        pid = state_data.get("promo_id")
        await db.update_promo(pid, max_per_user=val)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await bot.send_message(chat_id, "✅ Лимит на человека обновлён.")
        await _send_promo_detail(bot, chat_id, promo)
        return True

    if state == "adm_promo_edit_group":
        pid = state_data.get("promo_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_promo(pid, allowed_group=val)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await bot.send_message(chat_id, "✅ Группа разрешённых обновлена.")
        await _send_promo_detail(bot, chat_id, promo)
        return True

    if state == "adm_promo_edit_expiry":
        pid = state_data.get("promo_id")
        if text.strip() == "/empty":
            await db.update_promo(pid, expires_at=None)
        else:
            dt = _parse_datetime(text.strip())
            if not dt:
                await bot.send_message(chat_id, "Формат: ДД.ММ.ГГГГ ЧЧ:ММ")
                return True
            await db.update_promo(pid, expires_at=dt)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await bot.send_message(chat_id, "✅ Срок действия обновлён.")
        await _send_promo_detail(bot, chat_id, promo)
        return True

    if state == "adm_pay_name":
        provider = state_data.get("provider", "")
        set_state(user_id, "adm_pay_shop_id", provider=provider, pay_name=text)
        await bot.send_message(
            chat_id,
            f"Название: **{text}**\n\nВведите Shop ID (идентификатор магазина):",
            keyboard=akb.admin_payment_cancel(),
        )
        return True

    if state == "adm_pay_shop_id":
        provider = state_data.get("provider", "")
        pay_name = state_data.get("pay_name", "")
        set_state(user_id, "adm_pay_secret", provider=provider,
                  pay_name=pay_name, shop_id=text)
        await bot.send_message(
            chat_id,
            f"Название: **{pay_name}**\nShop ID: **{text}**\n\n"
            "Введите секретный ключ (Secret Key):",
            keyboard=akb.admin_payment_cancel(),
        )
        return True

    if state == "adm_pay_secret":
        provider = state_data.get("provider", "")
        pay_name = state_data.get("pay_name", "")
        shop_id = state_data.get("shop_id", "")
        method = await db.create_payment_method(pay_name, provider, shop_id, text)
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"✅ Способ оплаты «{pay_name}» создан!\n"
            f"Провайдер: {provider}\nShop ID: {shop_id}",
            keyboard=akb.admin_payment_detail(method["id"], method["is_active"]),
        )
        return True

    return False



async def _show_tariff_settings(reply_fn, tariff: dict):
    desc = tariff["description"] or "(пусто)"
    if len(desc) > 200:
        desc = desc[:200] + "…"
    await reply_fn(
        f"Настройка тарифа «{tariff['name']}»\n"
        f"Описание тарифа:\n{desc}",
        keyboard=akb.admin_tariff_settings(tariff["id"], tariff["is_active"]),
    )


async def _send_tariff_settings(bot: MaxBot, chat_id: int, tariff: dict):
    desc = tariff["description"] or "(пусто)"
    if len(desc) > 200:
        desc = desc[:200] + "…"
    await bot.send_message(
        chat_id,
        f"Настройка тарифа «{tariff['name']}»\n"
        f"Описание тарифа:\n{desc}",
        keyboard=akb.admin_tariff_settings(tariff["id"], tariff["is_active"]),
    )


def _parse_nullable_float(text: str) -> float | None:
    if text.strip() == "/empty":
        return None
    try:
        return float(text.replace(",", ".").replace(" ", ""))
    except ValueError:
        return None


def _parse_datetime(text: str) -> datetime | None:
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


async def _send_user_profile(bot: MaxBot, chat_id: int, profile: dict):
    uid = profile["user_id"]
    name = f"{profile['first_name']} {profile['last_name']}".strip() or f"id:{profile['user_id']}"

    subs_lines = []
    for p in profile["purchases"]:
        date_str = p["purchased_at"].strftime("%d.%m.%Y") if p.get("purchased_at") else "?"
        subs_lines.append(f" • {p['tariff_name']} — {date_str}")
    subs_text = "\n".join(subs_lines) if subs_lines else " нет покупок"

    text = (
        f"User ID: **{uid}**\n"
        f"Имя пользователя: {user_link(name, uid)}\n\n"
        f"Подписки:\n{subs_text}\n\n"
        f"Кол-во оплат: **{profile['total_count']}**\n"
        f"Сумма оплат: **{profile['total_paid']:.0f}₽**\n"
        f"Средний чек: **{profile['avg_check']:.0f}₽**"
    )
    await bot.send_message(chat_id, text, keyboard=akb.admin_user_profile(uid), fmt="markdown")


async def _send_user_profile_with_logs(bot: MaxBot, chat_id: int, profile: dict):
    uid = profile["user_id"]
    name = f"{profile['first_name']} {profile['last_name']}".strip() or f"id:{profile['user_id']}"

    active_tariffs = []
    for p in profile["purchases"]:
        if p.get("status") == "active":
            exp = p.get("expires_at")
            if exp is None or exp > datetime.now():
                active_tariffs.append(p["tariff_name"])
    active_str = ", ".join(active_tariffs) if active_tariffs else "нет"

    logs = await db.get_user_logs(uid)
    if logs:
        from collections import defaultdict
        by_date = defaultdict(list)
        for log in reversed(logs):
            date_str = log["created_at"].strftime("%d.%m.%Y")
            time_str = log["created_at"].strftime("%H:%M")
            by_date[date_str].append(f"{time_str}: {log['action']}")
        logs_lines = []
        for date, actions in by_date.items():
            logs_lines.append(f"\n**{date}**")
            for a in actions:
                logs_lines.append(a)
        logs_text = "\n".join(logs_lines)
    else:
        logs_text = "нет логов"

    text = (
        f"Пользователь: {user_link(name, uid)}\n"
        f"ID: **{uid}**\n"
        f"Активные тарифы: {active_str}\n\n"
        f"Логи:{logs_text}"
    )
    await bot.send_message(chat_id, text, keyboard=akb.admin_user_profile(uid), fmt="markdown")


async def _send_users_xlsx(bot: MaxBot, chat_id: int):
    from openpyxl import Workbook

    users = await db.all_users_with_purchases()
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    ws.append([
        "Имя Фамилия", "User ID", "Номер телефона",
        "Кол-во покупок", "Общая сумма оплат",
        "Купленные тарифы (✅ активен / ❌ нет)",
    ])
    for u in users:
        full_name = f"{u['first_name']} {u['last_name']}".strip()
        ws.append([
            full_name,
            u["user_id"],
            u.get("phone") or "",
            u["purchase_count"],
            float(u["total_paid"]),
            u["purchases"] or "",
        ])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        await bot.send_file(chat_id, tmp.name, "users.xlsx")
    finally:
        os.unlink(tmp.name)


async def _send_subscribers_xlsx(bot: MaxBot, chat_id: int):
    from openpyxl import Workbook

    rows = await db.tariff_subscribers()
    wb = Workbook()
    sheets_created = set()
    default_removed = False

    for r in rows:
        tariff_name = r["tariff_name"] or "Без названия"
        sheet_name = tariff_name[:31]
        if sheet_name not in sheets_created:
            if not default_removed:
                ws = wb.active
                ws.title = sheet_name
                default_removed = True
            else:
                ws = wb.create_sheet(title=sheet_name)
            ws.append([
                "Имя Фамилия", "User ID", "Телефон",
                "Дата покупки", "Время покупки", "Действует до", "Сумма оплаты",
            ])
            sheets_created.add(sheet_name)
        else:
            ws = wb[sheet_name]

        full_name = f"{r['first_name']} {r['last_name']}".strip()
        purchased_dt = r.get("purchased_at")
        purchased_date = purchased_dt.strftime("%d.%m.%Y") if purchased_dt else ""
        purchased_time = purchased_dt.strftime("%H:%M") if purchased_dt else ""
        expires = r["expires_at"].strftime("%d.%m.%Y") if r.get("expires_at") else "бессрочно"
        price_paid = r.get("price_paid")
        price_str = f"{price_paid:.0f}₽" if price_paid else "бесплатно"
        ws.append([full_name, r["user_id"], r.get("phone") or "", purchased_date, purchased_time, expires, price_str])

    if not rows:
        ws = wb.active
        ws.title = "Нет подписчиков"
        ws.append(["Активных подписчиков нет"])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        await bot.send_file(chat_id, tmp.name, "subscribers.xlsx")
    finally:
        os.unlink(tmp.name)


async def _send_expired_xlsx(bot: MaxBot, chat_id: int):
    from openpyxl import Workbook

    rows = await db.tariff_expired_subscribers()
    wb = Workbook()
    sheets_created = set()
    default_removed = False

    for r in rows:
        tariff_name = r["tariff_name"] or "Без названия"
        sheet_name = tariff_name[:31]
        if sheet_name not in sheets_created:
            if not default_removed:
                ws = wb.active
                ws.title = sheet_name
                default_removed = True
            else:
                ws = wb.create_sheet(title=sheet_name)
            ws.append([
                "Имя Фамилия", "User ID", "Телефон",
                "Дата покупки", "Истекла",
            ])
            sheets_created.add(sheet_name)
        else:
            ws = wb[sheet_name]

        full_name = f"{r['first_name']} {r['last_name']}".strip()
        purchased = r["purchased_at"].strftime("%d.%m.%Y") if r.get("purchased_at") else ""
        expired = r["expires_at"].strftime("%d.%m.%Y") if r.get("expires_at") else ""
        ws.append([full_name, r["user_id"], r.get("phone") or "", purchased, expired])

    if not rows:
        ws = wb.active
        ws.title = "Нет данных"
        ws.append(["Не продливших подписку нет"])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        await bot.send_file(chat_id, tmp.name, "expired.xlsx")
    finally:
        os.unlink(tmp.name)



async def _show_promo_detail(reply_fn, promo: dict):
    max_act = promo["max_activations"]
    max_act_str = "безлимит" if max_act == 0 else str(max_act)
    expires = promo["expires_at"].strftime("%d.%m.%Y %H:%M") if promo.get("expires_at") else "не ограничен"
    used = await db.count_promo_activations(promo["id"])
    text = (
        f"Промокод — **{promo['code']}**\n"
        f"Процент скидки — **{promo['discount_percent']}%**\n"
        f"Максимум активаций: **{max_act_str}**\n"
        f"Максимум активаций одним человеком — **{promo['max_per_user']}**\n"
        f"Срок действия: **{expires}**\n"
        f"Использовано: **{used}**"
    )
    await reply_fn(text, keyboard=akb.admin_promo_detail(promo["id"]))


async def _send_promo_detail(bot: MaxBot, chat_id: int, promo: dict):
    max_act = promo["max_activations"]
    max_act_str = "безлимит" if max_act == 0 else str(max_act)
    expires = promo["expires_at"].strftime("%d.%m.%Y %H:%M") if promo.get("expires_at") else "не ограничен"
    used = await db.count_promo_activations(promo["id"])
    text = (
        f"Промокод — **{promo['code']}**\n"
        f"Процент скидки — **{promo['discount_percent']}%**\n"
        f"Максимум активаций: **{max_act_str}**\n"
        f"Максимум активаций одним человеком — **{promo['max_per_user']}**\n"
        f"Срок действия: **{expires}**\n"
        f"Использовано: **{used}**"
    )
    await bot.send_message(chat_id, text, keyboard=akb.admin_promo_detail(promo["id"]))


async def _send_promo_activations_xlsx(bot: MaxBot, chat_id: int, promo_id: int):
    from openpyxl import Workbook

    activations = await db.get_promo_activations(promo_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Активации"
    ws.append(["ID", "Имя Фамилия", "Оплатил"])
    for a in activations:
        full_name = f"{a.get('first_name', '')} {a.get('last_name', '')}".strip()
        paid_str = "Да" if a.get("paid") else "Нет"
        ws.append([a["user_id"], full_name, paid_str])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        await bot.send_file(chat_id, tmp.name, "promo_activations.xlsx")
    finally:
        os.unlink(tmp.name)



def _contact_request_kb() -> dict:
    return {
        "type": "inline_keyboard",
        "payload": {"buttons": [
            [{"type": "request_contact", "text": "📱 Отправить номер телефона"}],
        ]},
    }


def _parse_broadcast_button_lines(text: str) -> tuple[list[dict], str | None]:
    return parse_inline_button_lines(text)


def _format_broadcast_buttons_message(buttons: list[dict]) -> str:
    return format_inline_buttons_message(buttons)


def _build_broadcast_keyboard(buttons: list[dict]) -> dict:
    return build_inline_keyboard(buttons)


async def _render_broadcast_text(text: str, user_id: int) -> str:
    user = await db.get_user(user_id)
    return format_template(text, **build_user_template_context(user, fallback=str(user_id)))
