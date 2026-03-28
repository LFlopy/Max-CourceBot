"""Обработчики админ-панели."""

import os
import tempfile
from datetime import datetime
import re
from max_client import MaxBot
from config import ADMIN_IDS
import database as db
import admin_keyboards as akb

from fsm import set_state, get_state, clear_state, user_states


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def _parse_duration_to_minutes(text: str) -> int | None:
    """
    Парсит длительность в минутах из строк вида:
    - "10м", "120m", "2ч", "48h"
    - "10 минут", "2 часа"
    Возвращает minutes или None, если распарсить нельзя.
    """
    s = (text or "").strip().lower()
    if not s:
        return None

    m = re.search(r"(\d+)\s*(ч|час|часа|часов|h)\b", s)
    if m:
        return int(m.group(1)) * 60

    m = re.search(r"(\d+)\s*(м|мин|минута|минуты|минут|m)\b", s)
    if m:
        return int(m.group(1))

    # если просто число — считаем минутами
    if re.fullmatch(r"\d+", s):
        return int(s)

    return None


# ── Флаг «показывать цену в названии» (in-memory) ────────────
_show_price_in_name = False


async def handle_admin_callback(bot: MaxBot, update: dict) -> bool:
    """Обрабатывает callback-и с префиксом 'adm:'.
    Возвращает True если обработал, False — если не админский callback.
    """
    callback = update.get("callback", update)
    callback_id = callback.get("callback_id", "")
    payload = callback.get("payload", "")
    sender = callback.get("user", callback.get("sender", {}))
    user_id = int(sender.get("user_id", 0))

    if not payload.startswith("adm:"):
        return False

    if not is_admin(user_id):
        await bot.answer_callback(callback_id, text="⛔ Нет доступа")
        return True

    msg = callback.get("message", {})
    body = msg.get("body", {})
    message_id = body.get("mid", "")
    recipient = msg.get("recipient", {}) if msg else {}
    chat_id = int(recipient.get("chat_id") or user_id)

    await bot.answer_callback(callback_id)

    async def reply(text: str, keyboard=None):
        ok = await bot.edit_message(message_id, text, keyboard=keyboard)
        if not ok:
            await bot.send_message(chat_id, text, keyboard=keyboard)

    global _show_price_in_name

    # ── Главное меню админки ──────────────────────────────
    if payload == "adm:back" or payload == "adm:menu":
        clear_state(user_id)
        await reply("Административная панель", keyboard=akb.admin_main())

    # ── Статистика ───────────────────────────────────────
    elif payload == "adm:stats":
        s = await db.stats_summary()
        text = (
            "📊 Статистика\n\n"
            f"👤 Всего пользователей: {s['total_users']}\n"
            f"🆕 Новых за 30 дней: {s['new_users_30d']}\n\n"
            f"🛒 Активных покупок: {s['total_purchases']}\n"
            f"🛒 Покупок за 30 дней: {s['purchases_30d']}\n"
            f"💰 Выручка за 30 дней: {s['revenue_30d']:.0f}₽"
        )
        await reply(text, keyboard=akb.admin_main())

    # ── Список тарифов ────────────────────────────────────
    elif payload == "adm:tariffs":
        clear_state(user_id)
        tariffs = await db.list_tariffs()
        await reply(
            "Список ваших категорий и тарифов",
            keyboard=akb.admin_tariff_list(tariffs, show_price=_show_price_in_name),
        )

    # ── 🎁 Гифт файлы: выбор тарифов ───────────────────────
    elif payload == "adm:gifts":
        tariffs = await db.list_tariffs()
        selected: set[int] = set()
        set_state(user_id, "adm_gift_pick", selected_tariffs=selected)
        await reply(
            "Укажите тарифы, связанные с гифт файлом.",
            keyboard=akb.admin_gift_tariff_picker(tariffs, selected),
        )

    elif payload.startswith("adm:gift_toggle:"):
        tid = int(payload.split(":")[2])
        state_data = user_states.get(user_id, {})
        selected: set[int] = state_data.get("selected_tariffs", set())
        if tid in selected:
            selected.discard(tid)
        else:
            selected.add(tid)
        state_data["selected_tariffs"] = selected
        tariffs = await db.list_tariffs()
        await reply(
            "Укажите тарифы, связанные с гифт файлом.",
            keyboard=akb.admin_gift_tariff_picker(tariffs, selected),
        )

    elif payload == "adm:gift_next":
        state_data = user_states.get(user_id, {})
        selected: set[int] = state_data.get("selected_tariffs", set())
        # Разрешаем сохранить даже если пусто — тогда бонусов не будет выдано
        set_state(user_id, "adm_gift_wait_file", selected_tariffs=selected)
        await reply("Отправьте гифт файл", keyboard=akb.admin_gift_wait_file())

    # ── Переключить отображение цены ──────────────────────
    elif payload == "adm:toggle_price":
        _show_price_in_name = not _show_price_in_name
        tariffs = await db.list_tariffs()
        await reply(
            "Список ваших категорий и тарифов",
            keyboard=akb.admin_tariff_list(tariffs, show_price=_show_price_in_name),
        )

    # ── Режим изменения порядка ───────────────────────────
    elif payload == "adm:reorder":
        set_state(user_id, "adm_reorder")
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф для перемещения",
            keyboard=akb.admin_reorder_list(tariffs),
        )

    elif payload.startswith("adm:sel_reorder:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_reorder", selected_id=tid)
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф для перемещения",
            keyboard=akb.admin_reorder_list(tariffs, selected_id=tid),
        )

    elif payload.startswith("adm:move_up:"):
        tid = int(payload.split(":")[2])
        await db.move_tariff_up(tid)
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф для перемещения",
            keyboard=akb.admin_reorder_list(tariffs, selected_id=tid),
        )

    elif payload.startswith("adm:move_down:"):
        tid = int(payload.split(":")[2])
        await db.move_tariff_down(tid)
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф для перемещения",
            keyboard=akb.admin_reorder_list(tariffs, selected_id=tid),
        )

    # ── Добавить тариф: шаг 1 — название ─────────────────
    elif payload == "adm:add_tariff":
        set_state(user_id, "adm_create_name")
        await reply(
            "Добавление нового тарифа\n \nОтправьте боту название для тарифа:",
            keyboard=akb.admin_create_cancel(),
        )

    # ── Добавить категорию ────────────────────────────────
    elif payload == "adm:add_category":
        set_state(user_id, "adm_create_category")
        await reply(
            "Введите название новой категории:",
            keyboard=akb.admin_create_cancel(),
        )

    # ── Отмена создания ───────────────────────────────────
    elif payload == "adm:cancel_create":
        clear_state(user_id)
        tariffs = await db.list_tariffs()
        await reply(
            "Список ваших категорий и тарифов",
            keyboard=akb.admin_tariff_list(tariffs, show_price=_show_price_in_name),
        )

    # ── Создание: бесплатный ──────────────────────────────
    elif payload == "adm:create_free":
        state_data = user_states.get(user_id, {})
        name = state_data.get("tariff_name", "")
        set_state(user_id, "adm_create_resources", tariff_name=name, tariff_price=0, is_free=True)
        await reply(
            f"Название тарифа — {name}\n \nЦена: бесплатно\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )

    # ── Создание: выбор длительности ──────────────────────
    elif payload.startswith("adm:dur:"):
        dur_val = payload.split(":")[2]
        state_data = user_states.get(user_id, {})
        current_state = get_state(user_id)

        # ── Режим РЕДАКТИРОВАНИЯ существующего тарифа ──
        if current_state == "adm_edit_duration":
            tid = state_data.get("tariff_id")
            if dur_val == "custom":
                set_state(user_id, "adm_edit_duration_custom", tariff_id=tid)
                await reply(
                    "Введите срок действия в формате: число часов или минут\n"
                    "Например: 48ч или 120м",
                    keyboard=akb.admin_back_to_settings(tid),
                )
                return True
            days = int(dur_val)
            dur_text = "Навсегда" if days == 0 else f"{days} дн."
            duration_minutes = None
            await db.update_tariff(
                tid,
                duration_days=days if days > 0 else 0,
                duration_text=dur_text,
                duration_minutes=duration_minutes,
            )
            clear_state(user_id)
            tariff = await db.get_tariff(tid)
            await bot.send_message(chat_id, "Продолжительность обновлена ✅")
            await _send_tariff_settings(bot, chat_id, tariff)
            return True

        # ── Режим СОЗДАНИЯ нового тарифа ──
        name = state_data.get("tariff_name", "")
        price = state_data.get("tariff_price", 0)
        is_free = state_data.get("is_free", False)

        if dur_val == "custom":
            set_state(user_id, "adm_create_duration_custom",
                      tariff_name=name, tariff_price=price, is_free=is_free)
            await reply(
                "Введите срок действия в формате: число часов или минут\n"
                "Например: 48ч или 120м",
                keyboard=akb.admin_create_cancel(),
            )
            return True

        days = int(dur_val)
        dur_text = "Навсегда" if days == 0 else f"{days} дн."
        set_state(user_id, "adm_create_resources",
                  tariff_name=name, tariff_price=price, is_free=is_free,
                  duration_days=days if days > 0 else None, duration_text=dur_text)

        price_str = "бесплатно" if is_free else f"{price}₽"
        await reply(
            f"Название тарифа — {name}\n \n"
            f"Цена: {price_str}\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, "
            "нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )

    # ── Создание: перейти к ресурсам ──────────────────────
    elif payload == "adm:go_resources":
        state_data = user_states.get(user_id, {})
        chats = await bot.get_chats()
        set_state(user_id, "adm_create_pick_resources",
                  tariff_name=state_data.get("tariff_name", ""),
                  tariff_price=state_data.get("tariff_price", 0),
                  is_free=state_data.get("is_free", False),
                  duration_days=state_data.get("duration_days"),
                  duration_text=state_data.get("duration_text", ""),
                  chats=chats, selected_resources=set())

        name = state_data.get("tariff_name", "")
        price = state_data.get("tariff_price", 0)
        is_free = state_data.get("is_free", False)
        price_str = "бесплатно" if is_free else f"{price}₽"
        await reply(
            f"Тариф — {name}\nЦена: {price_str}\n \n"
            "Выберите ресурсы к которым нужно выдать доступ после покупки",
            keyboard=akb.admin_resource_picker(chats, set()),
        )

    # ── Создание / редактирование: переключить ресурс ─────
    elif payload.startswith("adm:res_toggle:"):
        res_id = int(payload.split(":")[2])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_resources", set())
        if res_id in selected:
            selected.discard(res_id)
        else:
            selected.add(res_id)
        state_data["selected_resources"] = selected

        chats = state_data.get("chats", [])
        is_edit = state_data.get("state") == "adm_edit_resources"
        edit_tid = state_data.get("tariff_id") if is_edit else None

        if is_edit:
            tariff = await db.get_tariff(edit_tid)
            header = f"Тариф — {tariff['name']}\n \n"
        else:
            name = state_data.get("tariff_name", "")
            price = state_data.get("tariff_price", 0)
            is_free = state_data.get("is_free", False)
            price_str = "бесплатно" if is_free else f"{price}₽"
            header = f"Тариф — {name}\nЦена: {price_str}\n \n"

        await reply(
            header + "Выберите ресурсы к которым нужно выдать доступ после покупки",
            keyboard=akb.admin_resource_picker(chats, selected, edit_tariff_id=edit_tid),
        )

    # ── Создание / редактирование: сохранить ресурсы ──────
    elif payload == "adm:res_save":
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_resources", set())
        chats = state_data.get("chats", [])
        is_edit = state_data.get("state") == "adm_edit_resources"

        # Собираем выбранные ресурсы
        resources = []
        for c in chats:
            cid = c.get("chat_id")
            if cid in selected:
                title = c.get("title", "")
                # Убираем маркер удалённого чата из заголовка
                if title.startswith("❓ ") and title.endswith(" (удалён)"):
                    title = title[2:].rsplit(" (удалён)", 1)[0]
                invite_link = c.get("link", "")
                resources.append({"chat_id": cid, "chat_title": title, "invite_link": invite_link})

        if is_edit:
            # Режим редактирования — обновляем ресурсы существующего тарифа
            tid = state_data.get("tariff_id")
            await db.set_tariff_resources(tid, resources)
            clear_state(user_id)
            tariff = await db.get_tariff(tid)
            await reply(
                f"✅ Ресурсы тарифа «{tariff['name']}» обновлены ({len(resources)} шт.)",
                keyboard=akb.admin_tariff_settings(tid, tariff["is_active"]),
            )
        else:
            # Режим создания — создаём тариф
            name = state_data.get("tariff_name", "")
            price = state_data.get("tariff_price", 0)
            is_free = state_data.get("is_free", False)
            duration_days = state_data.get("duration_days")
            duration_minutes = state_data.get("duration_minutes")
            duration_text = state_data.get("duration_text", "")

            tariff = await db.create_tariff(
                name=name, price=float(price), is_free=is_free,
                duration_days=duration_days,
                duration_minutes=duration_minutes,
                duration_text=duration_text,
            )
            if resources:
                await db.set_tariff_resources(tariff["id"], resources)

            clear_state(user_id)
            await reply(
                f"Тариф «{name}»\nУспешно создан ✅",
                keyboard=akb.admin_tariff_created(tariff["id"]),
            )

    # ── Просмотр тарифа (из списка) ──────────────────────
    elif payload.startswith("adm:tariff:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        if not tariff:
            await reply("Тариф не найден")
            return True
        await _show_tariff_settings(reply, tariff)

    # ── Настройки тарифа ──────────────────────────────────
    elif payload.startswith("adm:settings:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        if not tariff:
            await reply("Тариф не найден")
            return True
        await _show_tariff_settings(reply, tariff)

    # ── Сохранить настройки (просто вернуться) ────────────
    elif payload.startswith("adm:save_settings:"):
        tid = int(payload.split(":")[2])
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await reply(
            f"Настройки тарифа «{tariff['name']}» сохранены ✅",
            keyboard=akb.admin_tariff_settings(tid, tariff["is_active"]),
        )

    # ── Скрыть / показать тариф ───────────────────────────
    elif payload.startswith("adm:toggle_active:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        new_active = not tariff["is_active"]
        await db.update_tariff(tid, is_active=new_active)
        tariff = await db.get_tariff(tid)
        await _show_tariff_settings(reply, tariff)

    # ── Удалить тариф (подтверждение) ─────────────────────
    elif payload.startswith("adm:delete:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        await reply(
            f"⚠️ Вы уверены что хотите удалить тариф «{tariff['name']}»?",
            keyboard=akb.admin_confirm_delete(tid),
        )

    elif payload.startswith("adm:confirm_del:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        name = tariff["name"] if tariff else "?"
        await db.delete_tariff(tid)
        tariffs = await db.list_tariffs()
        await reply(
            f"Тариф «{name}» удалён 🗑",
            keyboard=akb.admin_tariff_list(tariffs, show_price=_show_price_in_name),
        )

    # ── Настройка: Название ───────────────────────────────
    elif payload.startswith("adm:set_name:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_name", tariff_id=tid)
        await reply(
            "Отправьте боту новое название тарифа:",
            keyboard=akb.admin_back_to_settings(tid),
        )

    # ── Настройка: Цена ───────────────────────────────────
    elif payload.startswith("adm:set_price:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        t = tariff
        await reply(
            f"Настройка цен тарифа «{t['name']}»\n \n"
            f"Стандартная цена: {t['price']}₽\n"
            f"Старая цена: {t['old_price'] or '-'}\n"
            f"Цена продления: {t['renewal_price'] or '-'}\n"
            f"Цена продления активной подписки: {t['active_renewal_price'] or '-'}",
            keyboard=akb.admin_price_settings(tid),
        )

    elif payload.startswith("adm:ep_std:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_price_std", tariff_id=tid)
        await reply("Введите новую стандартную цену:", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:ep_old:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_price_old", tariff_id=tid)
        await reply("Введите новую старую цену (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:ep_renew:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_price_renew", tariff_id=tid)
        await reply("Введите цену продления (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:ep_active:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_price_active", tariff_id=tid)
        await reply("Введите цену продления активной подписки (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    # ── Настройка: Описание ───────────────────────────────
    elif payload.startswith("adm:set_desc:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_desc", tariff_id=tid)
        await reply(
            "Отправьте боту новое описание тарифа\n(Или delete, чтобы удалить его):",
            keyboard=akb.admin_back_to_settings(tid),
        )

    # ── Настройка: Продолжительность ──────────────────────
    elif payload.startswith("adm:set_duration:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_duration", tariff_id=tid)
        tariff = await db.get_tariff(tid)
        await reply(
            f"Текущая продолжительность: {tariff['duration_text'] or tariff['duration_days'] or 'не задана'}\n \n"
            "Введите новый срок действия тарифа в днях "
            "или выберите из готовых вариантов ниже",
            keyboard=akb.admin_edit_duration(tid),
        )

    # ── Настройка: Даты ───────────────────────────────────
    elif payload.startswith("adm:set_dates:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        start_day = tariff["start_day"] or "не установлен"
        start_date = tariff["start_date"].strftime("%d.%m.%Y %H:%M") if tariff["start_date"] else "не установлен"
        end_date = tariff["end_date"].strftime("%d.%m.%Y %H:%M") if tariff["end_date"] else "не установлен"
        await reply(
            f"Тариф: {tariff['name']}\n \n"
            f"Дата начала тарифа: {start_date}\n"
            f"Дата конца тарифа: {end_date}",
            keyboard=akb.admin_date_settings(tid),
        )


    elif payload.startswith("adm:ed_start:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_start_date", tariff_id=tid)
        await reply(
            'Введите дату (в формате: 16.03.2026 07:54) или "/empty" чтобы сбросить её',
            keyboard=akb.admin_date_input_back(tid),
        )

    elif payload.startswith("adm:ed_end:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_end_date", tariff_id=tid)
        await reply(
            'Введите дату (в формате: 16.03.2026 07:54) или "/empty" чтобы сбросить её',
            keyboard=akb.admin_date_input_back(tid),
        )

    # ── Настройка: Ресурсы ────────────────────────────────
    elif payload.startswith("adm:set_resources:"):
        tid = int(payload.split(":")[2])
        chats = await bot.get_chats()
        existing = await db.get_tariff_resources(tid)
        selected = {r["chat_id"] for r in existing}
        # Карта сохранённых invite_link из БД
        existing_links = {r["chat_id"]: r.get("invite_link", "") for r in existing}
        # Объединяем: текущие чаты бота + сохранённые в БД (которых может уже нет в боте)
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        # Добавляем сохранённый invite_link к чатам из API (если в API нет link)
        for c in chats:
            cid = c.get("chat_id")
            if not c.get("link") and cid in existing_links and existing_links[cid]:
                c["link"] = existing_links[cid]
        for r in existing:
            if r["chat_id"] not in chat_ids_in_list:
                chats.append({
                    "chat_id": r["chat_id"],
                    "title": f"❓ {r.get('chat_title') or r['chat_id']} (удалён)",
                    "link": r.get("invite_link", ""),
                })
        set_state(user_id, "adm_edit_resources", tariff_id=tid, chats=chats, selected_resources=selected)
        tariff = await db.get_tariff(tid)
        await reply(
            f"Тариф — {tariff['name']}\n \n"
            "Выберите ресурсы к которым нужно выдать доступ после покупки",
            keyboard=akb.admin_resource_picker(chats, selected, edit_tariff_id=tid),
        )

    # ── Настройка: остальные поля (чек, интервал, текст успеха, лимит, группа) ─
    elif payload.startswith("adm:set_check:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_check_name", tariff_id=tid)
        await reply("Введите название в чеке (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_reject:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_reject", tariff_id=tid)
        await reply("Введите интервал отклонений в минутах (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_success:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_success", tariff_id=tid)
        await reply("Введите текст при успешной покупке (или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_limit:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_limit", tariff_id=tid)
        await reply("Введите лимит активаций (число, или /empty чтобы сбросить):", keyboard=akb.admin_back_to_settings(tid))

    elif payload.startswith("adm:set_allowed:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        tariffs = await db.list_tariffs()
        selected = set()
        if tariff.get("allowed_group"):
            selected = {int(x) for x in tariff["allowed_group"].split(",") if x.strip()}
        set_state(user_id, "adm_allowed_pick", tariff_id=tid, selected_allowed=selected)
        await reply(
            f"Тариф «{tariff['name']}»\n\n"
            "Выберите тарифы, подписчикам которых будет виден этот тариф.\n"
            "Если ни один не выбран — тариф виден всем.",
            keyboard=akb.admin_tariff_allowed_picker(tariffs, selected, tid),
        )

    elif payload.startswith("adm:allowed_toggle:"):
        parts = payload.split(":")
        tid = int(parts[2])
        toggle_tid = int(parts[3])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_allowed", set())
        if toggle_tid in selected:
            selected.discard(toggle_tid)
        else:
            selected.add(toggle_tid)
        state_data["selected_allowed"] = selected
        tariffs = await db.list_tariffs()
        tariff = await db.get_tariff(tid)
        await reply(
            f"Тариф «{tariff['name']}»\n\n"
            "Выберите тарифы, подписчикам которых будет виден этот тариф.\n"
            "Если ни один не выбран — тариф виден всем.",
            keyboard=akb.admin_tariff_allowed_picker(tariffs, selected, tid),
        )

    elif payload.startswith("adm:allowed_save:"):
        tid = int(payload.split(":")[2])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_allowed", set())
        val = ",".join(str(x) for x in selected) if selected else None
        await db.update_tariff(tid, allowed_group=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await reply("Группа разрешённых обновлена ✅")
        await _show_tariff_settings(reply, tariff)

    elif payload.startswith("adm:buy_link:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        if tariff["is_free"]:
            # Для бесплатных — кнопка активации
            btns = [
                [{"type": "callback", "text": "✅ Активировать тариф", "payload": f"activate:{tid}"}],
                [{"type": "callback", "text": "🔙 Назад", "payload": f"adm:settings:{tid}"}],
            ]
            await reply(
                f"Тариф «{tariff['name']}» — бесплатный.\n\n"
                "Нажмите кнопку ниже, чтобы активировать:",
                keyboard={"type": "inline_keyboard", "payload": {"buttons": btns}},
            )
        else:
            link = tariff.get("payment_link") or ""
            btns = []
            if link:
                btns.append([{"type": "link", "text": "💳 Перейти к оплате", "url": link}])
            btns.append([{"type": "callback", "text": "💳 Вызвать оплату в боте", "payload": f"pay:{tid}"}])
            btns.append([{"type": "callback", "text": "🔙 Назад", "payload": f"adm:settings:{tid}"}])
            text = f"Тариф «{tariff['name']}» — {tariff['price']}₽\n\n"
            if link:
                text += f"Ссылка на оплату: {link}\n\n"
            text += "Нажмите кнопку ниже для перехода к покупке:"
            await reply(
                text,
                keyboard={"type": "inline_keyboard", "payload": {"buttons": btns}},
            )

    elif payload.startswith("adm:set_category:"):
        tid = int(payload.split(":")[2])
        categories = await db.list_categories()
        tariff = await db.get_tariff(tid)
        btns = []
        for cat in categories:
            icon = "✅" if tariff["category_id"] == cat["id"] else "⬜"
            btns.append([{"type": "callback", "text": f"{icon} {cat['name']}", "payload": f"adm:pick_cat:{tid}:{cat['id']}"}])
        btns.append([{"type": "callback", "text": "❌ Без категории", "payload": f"adm:pick_cat:{tid}:0"}])
        btns.append([{"type": "callback", "text": "🔙 Назад", "payload": f"adm:settings:{tid}"}])
        await reply(
            f"Выберите категорию для тарифа «{tariff['name']}»:",
            keyboard={"type": "inline_keyboard", "payload": {"buttons": btns}},
        )

    elif payload.startswith("adm:pick_cat:"):
        parts = payload.split(":")
        tid = int(parts[2])
        cat_id = int(parts[3])
        await db.update_tariff(tid, category_id=cat_id if cat_id > 0 else None)
        tariff = await db.get_tariff(tid)
        await _show_tariff_settings(reply, tariff)

    elif payload.startswith("adm:set_media:"):
        tid = int(payload.split(":")[2])
        set_state(user_id, "adm_edit_media", tariff_id=tid)
        await reply("Отправьте ссылку на медиа (изображение) для тарифа\n(Или /empty чтобы удалить):", keyboard=akb.admin_back_to_settings(tid))

    # ══════════════════════════════════════════════════════════
    # ══  НАСТРОЙКИ БОТА  ═════════════════════════════════════
    # ══════════════════════════════════════════════════════════

    elif payload == "adm:settings_menu":
        clear_state(user_id)
        await reply(
            "Настройки вашего бота.\n"
            "Для того чтобы вернуться обратно в Главное меню, "
            "вы можете отправить команду /start",
            keyboard=akb.admin_bot_settings(),
        )

    # ── Управление ресурсами ─────────────────────────────────
    elif payload == "adm:manage_resources":
        clear_state(user_id)
        chats = await bot.get_chats()
        usage = await db.get_resource_usage()
        # Добавляем ресурсы из БД, которых нет в текущих чатах бота
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        for cid, tariff_names in usage.items():
            if cid not in chat_ids_in_list:
                chats.append({"chat_id": cid, "title": f"❓ {cid} (недоступен)"})
        await reply(
            "📋 Ресурсы бота (чаты/каналы).\n\n"
            "Нажмите на ресурс чтобы удалить его из бота и отвязать от всех тарифов.",
            keyboard=akb.admin_resources_list(chats, usage),
        )

    elif payload.startswith("adm:res_del:"):
        cid = int(payload.split(":")[2])
        chats = await bot.get_chats()
        usage = await db.get_resource_usage()
        # Находим название
        title = str(cid)
        for c in chats:
            if c.get("chat_id") == cid:
                title = c.get("title", str(cid))
                break
        tariff_names = usage.get(cid, [])
        text = f"Удалить ресурс **{title}**?\n\n"
        if tariff_names:
            text += "Используется в тарифах:\n"
            text += "\n".join(f" • {n}" for n in tariff_names)
            text += "\n\nРесурс будет отвязан от всех тарифов."
        else:
            text += "Не привязан ни к одному тарифу."
        text += "\nБот покинет этот чат."
        await reply(text, keyboard=akb.admin_confirm_res_delete(cid))

    elif payload.startswith("adm:res_del_confirm:"):
        cid = int(payload.split(":")[2])
        # Отвязываем от тарифов
        await db.delete_resource_from_all_tariffs(cid)
        # Бот покидает чат
        await bot.leave_chat(cid)
        # Обновляем список
        chats = await bot.get_chats()
        usage = await db.get_resource_usage()
        chat_ids_in_list = {c.get("chat_id") for c in chats}
        for c_id, _ in usage.items():
            if c_id not in chat_ids_in_list:
                chats.append({"chat_id": c_id, "title": f"❓ {c_id} (недоступен)"})
        await reply(
            "✅ Ресурс удалён.\n\n"
            "📋 Ресурсы бота (чаты/каналы).",
            keyboard=akb.admin_resources_list(chats, usage),
        )

    # ── Подписчики ──────────────────────────────────────────
    elif payload == "adm:subscribers":
        clear_state(user_id)
        s = await db.subscribers_stats()
        text = (
            f"Всего переходов в бота: **{s['total_users']}** чел.\n\n"
            "Количество подписчиков:\n"
            f" • Активных подписок: **{s['active_subs']}** чел.\n"
            f" • Не продлили подписку: **{s['expired']}** чел.\n"
            f" • Купили подписку: **{s['bought']}** чел.\n"
            f" • Ни разу не купивших: **{s['never_bought']}** чел."
        )
        await reply(text, keyboard=akb.admin_subscribers())

    # ── Поиск по ID ─────────────────────────────────────────
    elif payload == "adm:sub_search_id":
        set_state(user_id, "adm_search_id_input")
        await reply(
            "Введите ID пользователя:",
            keyboard=akb.admin_back_subscribers(),
        )

    # ── Профиль пользователя ────────────────────────────────
    elif payload == "adm:sub_profile":
        set_state(user_id, "adm_profile_input")
        await reply(
            "Введите id пользователя, чтобы получить их профиль:",
            keyboard=akb.admin_back_subscribers(),
        )

    # ── Список подписчиков (.xlsx) ──────────────────────────
    elif payload == "adm:sub_list":
        try:
            await _send_subscribers_xlsx(bot, chat_id)
            await bot.send_message(chat_id, "👆 Список подписчиков", keyboard=akb.admin_subscribers())
        except Exception as e:
            await bot.send_message(chat_id, f"❌ Ошибка: {e}", keyboard=akb.admin_subscribers())

    # ── Список не продливших (.xlsx) ─────────────────────────
    elif payload == "adm:sub_expired":
        try:
            await _send_expired_xlsx(bot, chat_id)
            await bot.send_message(chat_id, "👆 Список не продливших", keyboard=akb.admin_subscribers())
        except Exception as e:
            await bot.send_message(chat_id, f"❌ Ошибка: {e}", keyboard=akb.admin_subscribers())

    # ── Таблица пользователей (.xlsx) ───────────────────────
    elif payload == "adm:sub_table":
        try:
            await _send_users_xlsx(bot, chat_id)
            await bot.send_message(chat_id, "👆 Таблица пользователей", keyboard=akb.admin_subscribers())
        except Exception as e:
            await bot.send_message(chat_id, f"❌ Ошибка: {e}", keyboard=akb.admin_subscribers())

    # ── Выдать подписку ─────────────────────────────────────
    elif payload == "adm:sub_grant":
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите какую подписку вы хотите подарить пользователю",
            keyboard=akb.admin_grant_tariff_list(tariffs),
        )

    elif payload.startswith("adm:grant_pick:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        set_state(user_id, "adm_grant_user", tariff_id=tid)
        await reply(
            f"Отправьте user\\_ID пользователя которому вы хотите "
            f"выдать подписку «{tariff['name']}».\n\n"
            "**!Обратите внимание, что для выдачи подписки, "
            "пользователь, которому активируем подписку, "
            "должен запустить бота, написав /start**",
            keyboard=akb.admin_back_subscribers(),
        )

    # ── Обнулить подписку ───────────────────────────────────
    elif payload == "adm:sub_revoke":
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите подписку для обнуления",
            keyboard=akb.admin_revoke_tariff_list(tariffs),
        )

    elif payload.startswith("adm:revoke_pick:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        set_state(user_id, "adm_revoke_user", tariff_id=tid)
        await reply(
            f"Отправьте user\\_ID пользователя, у которого нужно "
            f"обнулить подписку «{tariff['name']}».",
            keyboard=akb.admin_back_subscribers(),
        )

    # ── Передать подписку ───────────────────────────────────
    elif payload == "adm:sub_transfer":
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите подписку для передачи",
            keyboard=akb.admin_transfer_tariff_list(tariffs),
        )

    elif payload.startswith("adm:transfer_pick:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        set_state(user_id, "adm_transfer_from", tariff_id=tid)
        await reply(
            f"Отправьте user\\_ID пользователя, **от которого** "
            f"передать подписку «{tariff['name']}».",
            keyboard=akb.admin_back_subscribers(),
        )

    # ── Написать пользователю ───────────────────────────────
    elif payload.startswith("adm:msg_user:"):
        target = int(payload.split(":")[2])
        set_state(user_id, "adm_msg_user", target_user_id=target)
        await reply(
            f"Введите сообщение для пользователя (id={target}):",
            keyboard=akb.admin_back_subscribers(),
        )

    # ── Таблица платежей пользователя ─────────────────────────
    elif payload.startswith("adm:pay_table:"):
        target = int(payload.split(":")[2])
        profile = await db.user_profile(target)
        if not profile:
            await reply("Пользователь не найден.")
            return
        purchases = profile["purchases"]
        if not purchases:
            await reply("У пользователя нет платежей.",
                        keyboard=akb.admin_user_profile(target))
            return
        lines = [f"🧾 Платежи пользователя {target}:\n"]
        for p in purchases:
            date_str = p["purchased_at"].strftime("%d.%m.%Y %H:%M") if p.get("purchased_at") else "?"
            status_map = {"active": "✅", "pending": "⏳", "expired": "⌛", "canceled": "❌", "revoked": "🚫"}
            icon = status_map.get(p["status"], "❓")
            price = f"{p['price_paid']}₽" if p.get("price_paid") else "бесплатно"
            lines.append(f"{icon} {p['tariff_name']} — {price} — {date_str} ({p['status']})")
        await reply("\n".join(lines), keyboard=akb.admin_user_profile(target))

    # ── Таблица подписок пользователя ──────────────────────────
    elif payload.startswith("adm:sub_table_user:"):
        target = int(payload.split(":")[2])
        profile = await db.user_profile(target)
        if not profile:
            await reply("Пользователь не найден.")
            return
        purchases = profile["purchases"]
        active = [p for p in purchases if p["status"] == "active"]
        if not active:
            await reply("У пользователя нет активных подписок.",
                        keyboard=akb.admin_user_profile(target))
            return
        lines = [f"📋 Подписки пользователя {target}:\n"]
        for p in active:
            exp = p.get("expires_at")
            if exp:
                exp_str = exp.strftime("%d.%m.%Y %H:%M")
            else:
                exp_str = "бессрочно"
            activated = p["activated_at"].strftime("%d.%m.%Y") if p.get("activated_at") else "?"
            lines.append(f"• {p['tariff_name']} — с {activated} до {exp_str}")
        await reply("\n".join(lines), keyboard=akb.admin_user_profile(target))

    # ══════════════════════════════════════════════════════════
    # ══  РАССЫЛКА  ════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════

    elif payload == "adm:broadcast":
        clear_state(user_id)
        await reply(
            "Выберите группу пользователей для рассылки:",
            keyboard=akb.admin_broadcast_groups(),
        )

    elif payload == "adm:bc_all":
        set_state(user_id, "adm_broadcast", bc_group="all")
        await reply(
            "Группа: **Все пользователи**\n\nОтправьте текст рассылки:",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_paid":
        set_state(user_id, "adm_broadcast", bc_group="paid")
        await reply(
            "Группа: **Оплатили тариф**\n\nОтправьте текст рассылки:",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_no_sub":
        set_state(user_id, "adm_broadcast", bc_group="no_sub")
        await reply(
            "Группа: **Без подписки**\n\nОтправьте текст рассылки:",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_pending":
        set_state(user_id, "adm_broadcast", bc_group="pending")
        await reply(
            "Группа: **Вызвал оплату, но не оплатил**\n\nОтправьте текст рассылки:",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_tariff":
        tariffs = await db.list_tariffs()
        await reply(
            "Выберите тариф, подписчикам которого нужно отправить рассылку:",
            keyboard=akb.admin_broadcast_tariff_list(tariffs),
        )

    elif payload.startswith("adm:bc_tariff_pick:"):
        tid = int(payload.split(":")[2])
        tariff = await db.get_tariff(tid)
        set_state(user_id, "adm_broadcast", bc_group="tariff", bc_tariff_id=tid)
        await reply(
            f"Группа: **подписчики «{tariff['name']}»**\n\nОтправьте текст рассылки:",
            keyboard=akb.admin_broadcast_cancel(),
        )

    elif payload == "adm:bc_btn_none":
        sd = user_states.get(user_id, {})
        bc_group = sd.get("bc_group", "all")
        bc_tariff_id = sd.get("bc_tariff_id")
        bc_text = sd.get("bc_text", "")
        clear_state(user_id)
        if bc_group == "all":
            user_ids = await db.get_all_user_ids()
        elif bc_group == "paid":
            user_ids = await db.get_paid_user_ids()
        elif bc_group == "no_sub":
            user_ids = await db.get_no_sub_user_ids()
        elif bc_group == "pending":
            user_ids = await db.get_pending_user_ids()
        elif bc_group == "tariff" and bc_tariff_id:
            user_ids = await db.get_tariff_user_ids(bc_tariff_id)
        else:
            user_ids = []
        sent = 0
        for uid in user_ids:
            try:
                await bot.send_message(uid, bc_text)
                sent += 1
            except Exception:
                pass
        await reply(
            f"✅ Рассылка завершена.\nОтправлено: **{sent}** из **{len(user_ids)}** пользователей.",
            keyboard=akb.admin_bot_settings(),
        )

    elif payload.startswith("adm:bc_btn_tariff:"):
        tid = int(payload.split(":")[2])
        sd = user_states.get(user_id, {})
        bc_group = sd.get("bc_group", "all")
        bc_tariff_id = sd.get("bc_tariff_id")
        bc_text = sd.get("bc_text", "")
        tariff = await db.get_tariff(tid)
        clear_state(user_id)
        if bc_group == "all":
            user_ids = await db.get_all_user_ids()
        elif bc_group == "paid":
            user_ids = await db.get_paid_user_ids()
        elif bc_group == "no_sub":
            user_ids = await db.get_no_sub_user_ids()
        elif bc_group == "pending":
            user_ids = await db.get_pending_user_ids()
        elif bc_group == "tariff" and bc_tariff_id:
            user_ids = await db.get_tariff_user_ids(bc_tariff_id)
        else:
            user_ids = []
        btn_label = f"💳 Оформить «{tariff['name']}»"
        tariff_btn = {"type": "callback", "text": btn_label, "payload": f"pay:{tid}"}
        bc_keyboard = {"type": "inline_keyboard", "payload": {"buttons": [[tariff_btn]]}}
        sent = 0
        for uid in user_ids:
            try:
                await bot.send_message(uid, bc_text, keyboard=bc_keyboard)
                sent += 1
            except Exception:
                pass
        await reply(
            f"✅ Рассылка завершена.\nОтправлено: **{sent}** из **{len(user_ids)}** пользователей.",
            keyboard=akb.admin_bot_settings(),
        )

    # ── Сбор контактов ─────────────────────────────────────────
    elif payload == "adm:collect_contacts":
        user_ids = await db.get_all_user_ids()
        sent = 0
        for uid in user_ids:
            try:
                await bot.send_message(
                    uid,
                    "📱 Пожалуйста, поделитесь вашим номером телефона, "
                    "нажав кнопку ниже 👇",
                    keyboard=_contact_request_kb(),
                )
                sent += 1
            except Exception:
                pass
        await reply(
            f"✅ Запрос на получение номера телефона отправлен {sent} пользователям.",
            keyboard=akb.admin_bot_settings(),
        )

    # ── Ответить на обратную связь ────────────────────────────
    elif payload.startswith("adm:reply_feedback:"):
        target = int(payload.split(":")[2])
        set_state(user_id, "adm_reply_feedback", target_user_id=target)
        await reply(
            "Ответьте на вопрос. Это может быть текст, фото, видео "
            "или любое другое медиа вложение:",
            keyboard=akb.admin_cancel_feedback_reply(),
        )

    # ── Отмена ответа на обратную связь ─────────────────────────
    elif payload == "adm:cancel_feedback_reply":
        clear_state(user_id)
        await reply("Ответ отменён.")

    # ── Забанить пользователя ─────────────────────────────────
    elif payload.startswith("adm:ban_user:"):
        target = int(payload.split(":")[2])
        await db.ban_user(target)
        target_user = await db.get_user(target)
        name = ""
        if target_user:
            name = f"{target_user['first_name']} {target_user['last_name']}".strip()
        await reply(
            f"🚫 Пользователь {name} (id={target}) заблокирован.",
        )

    # ══════════════════════════════════════════════════════════
    # ══  СПОСОБЫ ОПЛАТЫ  ═══════════════════════════════════
    # ══════════════════════════════════════════════════════════

    elif payload == "adm:payment_methods":
        clear_state(user_id)
        methods = await db.list_payment_methods()
        await reply(
            "💳 Способы оплаты.\n\nНажмите на метод для управления или добавьте новый.",
            keyboard=akb.admin_payment_methods_list(methods),
        )

    elif payload == "adm:add_pay_method":
        import payments as pay_mod
        providers = pay_mod.provider_names()
        await reply(
            "Выберите платёжную систему:",
            keyboard=akb.admin_payment_provider_list(providers),
        )

    elif payload.startswith("adm:pay_provider:"):
        provider_key = payload.split(":", 2)[2]
        set_state(user_id, "adm_pay_name", provider=provider_key)
        await reply(
            "Введите название способа оплаты\n(как он будет отображаться пользователям):",
            keyboard=akb.admin_payment_cancel(),
        )

    elif payload.startswith("adm:pay_detail:"):
        mid = int(payload.split(":")[2])
        method = await db.get_payment_method(mid)
        if not method:
            await reply("Метод не найден.", keyboard=akb.admin_payment_methods_list(await db.list_payment_methods()))
            return
        status = "✅ Включён" if method["is_active"] else "❌ Выключен"
        text = (
            f"💳 **{method['name']}**\n"
            f"Провайдер: {method['provider']}\n"
            f"Shop ID: {method['shop_id']}\n"
            f"Статус: {status}"
        )
        await reply(text, keyboard=akb.admin_payment_detail(mid, method["is_active"]))

    elif payload.startswith("adm:toggle_pay:"):
        mid = int(payload.split(":")[2])
        method = await db.toggle_payment_method(mid)
        if method:
            status = "включён ✅" if method["is_active"] else "выключен ❌"
            await reply(f"Метод «{method['name']}» — {status}",
                        keyboard=akb.admin_payment_detail(mid, method["is_active"]))

    elif payload.startswith("adm:del_pay:"):
        mid = int(payload.split(":")[2])
        method = await db.get_payment_method(mid)
        await reply(
            f"Удалить способ оплаты «{method['name']}»?",
            keyboard=akb.admin_confirm_pay_delete(mid),
        )

    elif payload.startswith("adm:del_pay_confirm:"):
        mid = int(payload.split(":")[2])
        await db.delete_payment_method(mid)
        methods = await db.list_payment_methods()
        await reply("✅ Способ оплаты удалён.", keyboard=akb.admin_payment_methods_list(methods))

    # ══════════════════════════════════════════════════════════
    # ══  РЕДАКТИРОВАНИЕ  ══════════════════════════════════════
    # ══════════════════════════════════════════════════════════

    elif payload == "adm:editing_menu":
        clear_state(user_id)
        await reply(
            "Выберите что вы хотите отредактировать:",
            keyboard=akb.admin_editing_menu(),
        )

    elif payload == "adm:button_texts":
        clear_state(user_id)
        await reply(
            "🔤 Редактирование текста кнопок Личного кабинета.\n"
            "Выберите кнопку для изменения:",
            keyboard=akb.admin_button_texts_list(db.BUTTON_TEXT_LABELS),
        )

    elif payload.startswith("adm:edit_btn:"):
        key = payload.split(":", 2)[2]
        label = db.BUTTON_TEXT_LABELS.get(key, key)
        current = await db.get_bot_text(key)
        set_state(user_id, "adm_edit_btn_text", text_key=key)
        await reply(
            f"**{label}**\n\nТекущий текст кнопки:\n{current}\n\n"
            "Отправьте новый текст:",
            keyboard=akb.admin_edit_btn_back(),
        )

    elif payload == "adm:desc_texts":
        clear_state(user_id)
        await reply(
            "📝 Редактирование описаний.\n"
            "Выберите описание для изменения:",
            keyboard=akb.admin_desc_texts_list(db.DESC_TEXT_LABELS),
        )

    elif payload.startswith("adm:edit_desc:"):
        key = payload.split(":", 2)[2]
        label = db.DESC_TEXT_LABELS.get(key, key)
        current = await db.get_bot_text(key)
        preview = current[:300] + "…" if len(current) > 300 else current
        set_state(user_id, "adm_edit_desc_text", text_key=key)
        await reply(
            f"**{label}**\n\nТекущий текст:\n{preview}\n\n"
            "Отправьте новый текст:",
            keyboard=akb.admin_edit_desc_back(),
        )

    # ══════════════════════════════════════════════════════════
    # ══  ОТВЕТЫ ОТ БОТА  ════════════════════════════════════
    # ══════════════════════════════════════════════════════════

    elif payload == "adm:bot_texts":
        clear_state(user_id)
        await reply(
            "Редактирование текстов бота.\nВыберите текст для изменения:",
            keyboard=akb.admin_bot_texts_list(db.BOT_TEXT_LABELS),
        )

    elif payload.startswith("adm:edit_text:"):
        key = payload.split(":", 2)[2]
        label = db.BOT_TEXT_LABELS.get(key, key)
        current = await db.get_bot_text(key)
        preview = current[:300] + "…" if len(current) > 300 else current
        set_state(user_id, "adm_edit_bot_text", text_key=key)
        await reply(
            f"**{label}**\n\nТекущий текст:\n{preview}\n\n"
            "Отправьте новый текст:",
            keyboard=akb.admin_bot_text_back(),
        )

    # ══════════════════════════════════════════════════════════
    # ══  ПРОМОКОДЫ  ══════════════════════════════════════════
    # ══════════════════════════════════════════════════════════

    elif payload == "adm:promo_menu":
        clear_state(user_id)
        await reply(
            "**Настройки промокодов**\n\n"
            "Выберите нужный раздел промокодов ниже или создайте новый:",
            keyboard=akb.admin_promo_menu(),
        )

    # ── Списки промокодов по типу ─────────────────────────────
    elif payload == "adm:promo_general":
        promos = await db.list_promos("general")
        if promos:
            await reply("Общие промокоды:", keyboard=akb.admin_promo_list(promos))
        else:
            set_state(user_id, "adm_promo_name", promo_type="general")
            await reply(
                "Отправьте боту новый промокод\n(например: discount30, PROMO20)",
                keyboard=akb.admin_promo_back(),
            )

    elif payload == "adm:promo_broadcast":
        promos = await db.list_promos("broadcast")
        if promos:
            await reply("Промокоды из рассылок:", keyboard=akb.admin_promo_list(promos))
        else:
            await reply("Промокодов из рассылок пока нет.", keyboard=akb.admin_promo_back())

    elif payload == "adm:promo_activation":
        promos = await db.list_promos("activation")
        if promos:
            await reply("Промокоды активационных ссылок:", keyboard=akb.admin_promo_list(promos))
        else:
            await reply("Промокодов активационных ссылок пока нет.", keyboard=akb.admin_promo_back())

    # ── Создать промокод ─────────────────────────────────────
    elif payload == "adm:promo_create":
        set_state(user_id, "adm_promo_name", promo_type="general")
        await reply(
            "Отправьте боту новый промокод\n(например: discount30, PROMO20)",
            keyboard=akb.admin_promo_back(),
        )

    # ── Создать группу промокодов ────────────────────────────
    elif payload == "adm:promo_create_group":
        set_state(user_id, "adm_promo_group_name")
        await reply(
            "Введите название группы промокодов:",
            keyboard=akb.admin_promo_back(),
        )

    # ── Открыть промокод ─────────────────────────────────────
    elif payload.startswith("adm:promo_open:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        if not promo:
            await reply("Промокод не найден.")
            return True
        await _show_promo_detail(reply, promo)

    # ── Разрешённые тарифы ───────────────────────────────────
    elif payload.startswith("adm:promo_tariffs:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        tariffs = await db.list_tariffs()
        selected = set()
        if promo.get("allowed_tariffs"):
            selected = {int(x) for x in promo["allowed_tariffs"].split(",") if x.strip()}
        set_state(user_id, "adm_promo_tariff_pick", promo_id=pid, selected_tariffs=selected)
        await reply(
            f"Промокод **{promo['code']}**\nВыберите тарифы, к которым можно применить промокод:",
            keyboard=akb.admin_promo_tariff_picker(tariffs, selected, pid),
        )

    elif payload.startswith("adm:promo_toggle_t:"):
        parts = payload.split(":")
        pid = int(parts[2])
        tid = int(parts[3])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_tariffs", set())
        if tid in selected:
            selected.discard(tid)
        else:
            selected.add(tid)
        state_data["selected_tariffs"] = selected
        tariffs = await db.list_tariffs()
        promo = await db.get_promo(pid)
        await reply(
            f"Промокод **{promo['code']}**\nВыберите тарифы:",
            keyboard=akb.admin_promo_tariff_picker(tariffs, selected, pid),
        )

    elif payload.startswith("adm:promo_save_t:"):
        pid = int(payload.split(":")[2])
        state_data = user_states.get(user_id, {})
        selected = state_data.get("selected_tariffs", set())
        val = ",".join(str(x) for x in selected) if selected else None
        await db.update_promo(pid, allowed_tariffs=val)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await reply("Разрешённые тарифы обновлены ✅")
        await _show_promo_detail(reply, promo)

    # ── Редактирование полей промокода ────────────────────────
    elif payload.startswith("adm:promo_edit_max:"):
        pid = int(payload.split(":")[2])
        set_state(user_id, "adm_promo_edit_max", promo_id=pid)
        await reply(
            "Введите максимальное кол-во активаций (0 = безлимит):",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_edit_per_user:"):
        pid = int(payload.split(":")[2])
        set_state(user_id, "adm_promo_edit_per_user", promo_id=pid)
        await reply(
            "Введите максимальное кол-во активаций на одного человека:",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_edit_group:"):
        pid = int(payload.split(":")[2])
        set_state(user_id, "adm_promo_edit_group", promo_id=pid)
        await reply(
            "Введите название группы разрешённых пользователей\n(или /empty чтобы сбросить):",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    elif payload.startswith("adm:promo_edit_expiry:"):
        pid = int(payload.split(":")[2])
        set_state(user_id, "adm_promo_edit_expiry", promo_id=pid)
        await reply(
            "Введите дату окончания в формате ДД.ММ.ГГГГ ЧЧ:ММ\n(или /empty для безлимитного срока):",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    # ── Список разрешённых пользователей ──────────────────────
    elif payload.startswith("adm:promo_allowed_users:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        users_str = promo.get("allowed_users") or "Не ограничено (все пользователи)"
        await reply(
            f"Промокод **{promo['code']}**\n\n"
            f"Разрешённые пользователи:\n{users_str}",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    # ── Список активаций (.xlsx) ──────────────────────────────
    elif payload.startswith("adm:promo_activations:"):
        pid = int(payload.split(":")[2])
        await _send_promo_activations_xlsx(bot, chat_id, pid)
        promo = await db.get_promo(pid)
        await bot.send_message(
            chat_id, f"👆 Активации промокода **{promo['code']}**",
            keyboard=akb.admin_promo_back_to_detail(pid),
        )

    # ── Удалить промокод ──────────────────────────────────────
    elif payload.startswith("adm:promo_delete:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        await reply(
            f"⚠️ Удалить промокод **{promo['code']}**?",
            keyboard=akb.admin_promo_confirm_delete(pid),
        )

    elif payload.startswith("adm:promo_confirm_del:"):
        pid = int(payload.split(":")[2])
        promo = await db.get_promo(pid)
        code = promo["code"] if promo else "?"
        await db.delete_promo(pid)
        await reply(
            f"Промокод **{code}** удалён 🗑",
            keyboard=akb.admin_promo_menu(),
        )

    else:
        await reply("Неизвестная команда админки")

    return True


async def handle_admin_message(
    bot: MaxBot,
    user_id: int,
    chat_id: int,
    text: str,
    attachments: list | None = None,
) -> bool:
    """Обрабатывает текстовые сообщения в контексте админ-FSM.
    Возвращает True если обработал.
    """
    state = get_state(user_id)
    if not state.startswith("adm_"):
        return False

    state_data = user_states.get(user_id, {})

    # ── Создание тарифа: ввод названия ────────────────────
    if state == "adm_create_name":
        set_state(user_id, "adm_create_price", tariff_name=text)
        await bot.send_message(
            chat_id,
            f"Название тарифа — {text}\n \nВведите цену или сделайте тариф бесплатным",
            keyboard=akb.admin_create_price(),
        )
        return True

    # ── Создание тарифа: ввод цены ────────────────────────
    if state == "adm_create_price":
        try:
            price = float(text.replace(",", ".").replace(" ", ""))
        except ValueError:
            await bot.send_message(chat_id, "Введите число (цену в рублях):", keyboard=akb.admin_create_price())
            return True
        name = state_data.get("tariff_name", "")
        set_state(user_id, "adm_create_resources", tariff_name=name, tariff_price=price, is_free=False)
        await bot.send_message(
            chat_id,
            f"Название тарифа — {name}\n \nЦена: {price}₽\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )
        return True

    # ── Создание тарифа: ввод произвольной длительности ───
    if state == "adm_create_duration_custom":
        name = state_data.get("tariff_name", "")
        price = state_data.get("tariff_price", 0)
        is_free = state_data.get("is_free", False)

        dur_text = text.strip()
        duration_minutes = _parse_duration_to_minutes(dur_text)
        set_state(user_id, "adm_create_resources",
                  tariff_name=name, tariff_price=price, is_free=is_free,
                  duration_days=None, duration_minutes=duration_minutes,
                  duration_text=dur_text)
        price_str = "бесплатно" if is_free else f"{price}₽"
        await bot.send_message(
            chat_id,
            f"Название тарифа — {name}\n \n"
            f"Цена: {price_str}\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, "
            "нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )
        return True

    # ── Создание тарифа: ввод длительности числом (дни) ───
    if state == "adm_create_duration":
        try:
            days = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите число дней:", keyboard=akb.admin_create_duration())
            return True
        name = state_data.get("tariff_name", "")
        price = state_data.get("tariff_price", 0)
        is_free = state_data.get("is_free", False)
        dur_text = f"{days} дн."
        set_state(user_id, "adm_create_resources",
                  tariff_name=name, tariff_price=price, is_free=is_free,
                  duration_days=days, duration_text=dur_text)
        price_str = "бесплатно" if is_free else f"{price}₽"
        await bot.send_message(
            chat_id,
            f"Название тарифа — {name}\n \n"
            f"Цена: {price_str}\n \n"
            "Чтобы перейти к добавлению каналов/групп к тарифу, "
            "нажмите кнопку ниже",
            keyboard=akb.admin_create_go_resources(),
        )
        return True

    # ── Создание категории ────────────────────────────────
    if state == "adm_create_category":
        await db.create_category(text.strip())
        clear_state(user_id)
        tariffs = await db.list_tariffs()
        await bot.send_message(
            chat_id,
            f"Категория «{text.strip()}» создана ✅\n \nСписок ваших категорий и тарифов",
            keyboard=akb.admin_tariff_list(tariffs),
        )
        return True

    # ── 🎁 Гифт файл: ожидание файла ───────────────────────
    if state == "adm_gift_wait_file":
        atts = attachments or []
        file_token = ""
        file_name = ""
        for att in atts:
            if not isinstance(att, dict):
                continue
            if att.get("type") != "file":
                continue
            payload = att.get("payload") or {}
            # максимально терпимо к формату
            file_token = payload.get("token") or payload.get("file_token") or ""
            file_name = payload.get("name") or payload.get("file_name") or ""
            if file_token:
                break

        if not file_token:
            await bot.send_message(chat_id, "Пришлите именно файл (вложение).")
            return True

        selected: set[int] = state_data.get("selected_tariffs", set())
        gift = await db.create_gift_file(file_token=file_token, file_name=file_name, tariff_ids=list(selected))
        clear_state(user_id)
        if gift:
            await bot.send_message(chat_id, "✅ Гифт файл сохранён.", keyboard=akb.admin_main())
        else:
            await bot.send_message(chat_id, "❌ Не удалось сохранить гифт файл.", keyboard=akb.admin_main())
        return True

    # ── Редактирование: название ──────────────────────────
    if state == "adm_edit_name":
        tid = state_data.get("tariff_id")
        await db.update_tariff(tid, name=text.strip())
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, f"Название обновлено ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: описание ──────────────────────────
    if state == "adm_edit_desc":
        tid = state_data.get("tariff_id")
        desc = "" if text.strip().lower() == "delete" else text.strip()
        await db.update_tariff(tid, description=desc)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Описание обновлено ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: стандартная цена ──────────────────
    if state == "adm_edit_price_std":
        tid = state_data.get("tariff_id")
        try:
            price = float(text.replace(",", ".").replace(" ", ""))
        except ValueError:
            await bot.send_message(chat_id, "Введите число:")
            return True
        await db.update_tariff(tid, price=price, is_free=(price == 0))
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Цена обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: старая цена ───────────────────────
    if state == "adm_edit_price_old":
        tid = state_data.get("tariff_id")
        val = _parse_nullable_float(text)
        await db.update_tariff(tid, old_price=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Старая цена обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: цена продления ────────────────────
    if state == "adm_edit_price_renew":
        tid = state_data.get("tariff_id")
        val = _parse_nullable_float(text)
        await db.update_tariff(tid, renewal_price=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Цена продления обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: цена продления активной ───────────
    if state == "adm_edit_price_active":
        tid = state_data.get("tariff_id")
        val = _parse_nullable_float(text)
        await db.update_tariff(tid, active_renewal_price=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Цена продления активной подписки обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: продолжительность ─────────────────
    if state == "adm_edit_duration":
        tid = state_data.get("tariff_id")
        try:
            days = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите число дней:")
            return True
        dur_text = f"{days} дн."
        await db.update_tariff(tid, duration_days=days, duration_text=dur_text, duration_minutes=None)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Продолжительность обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    if state == "adm_edit_duration_custom":
        tid = state_data.get("tariff_id")
        dur_text = text.strip()
        duration_minutes = _parse_duration_to_minutes(dur_text)
        if not duration_minutes:
            await bot.send_message(chat_id, "Не удалось распознать. Введите например: 48ч или 120м")
            return True
        await db.update_tariff(
            tid,
            duration_days=None,
            duration_minutes=duration_minutes,
            duration_text=dur_text,
        )
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Продолжительность обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: день начала ───────────────────────
    if state == "adm_edit_start_day":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, start_day=None)
        else:
            # Принимаем число (день) или дату (16.03.2026 / 16.03.2026 07:54)
            try:
                day = int(text.strip())
            except ValueError:
                dt = _parse_datetime(text.strip())
                if not dt:
                    await bot.send_message(chat_id, "Введите число (день) или дату (16.03.2026 07:54):")
                    return True
                day = dt.day
            await db.update_tariff(tid, start_day=day)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "День начала обновлён ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: дата начала ───────────────────────
    if state == "adm_edit_start_date":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, start_date=None)
        else:
            dt = _parse_datetime(text.strip())
            if not dt:
                await bot.send_message(chat_id, "Формат: 16.03.2026 07:54")
                return True
            await db.update_tariff(tid, start_date=dt)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Дата начала обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: дата конца ────────────────────────
    if state == "adm_edit_end_date":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, end_date=None)
        else:
            dt = _parse_datetime(text.strip())
            if not dt:
                await bot.send_message(chat_id, "Формат: 16.03.2026 07:54")
                return True
            await db.update_tariff(tid, end_date=dt)
            # Синхронизируем expires_at у всех активных подписок этого тарифа
            await db.update_active_purchases_expiry(tid, dt)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Дата конца обновлена ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: название в чеке ───────────────────
    if state == "adm_edit_check_name":
        tid = state_data.get("tariff_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_tariff(tid, check_name=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Название в чеке обновлено ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: интервал отклонений ───────────────
    if state == "adm_edit_reject":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, rejection_interval=None)
        else:
            try:
                val = int(text.strip())
            except ValueError:
                await bot.send_message(chat_id, "Введите число минут:")
                return True
            await db.update_tariff(tid, rejection_interval=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Интервал отклонений обновлён ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: текст успешной покупки ────────────
    if state == "adm_edit_success":
        tid = state_data.get("tariff_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_tariff(tid, success_text=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Текст успешной покупки обновлён ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: лимит активаций ───────────────────
    if state == "adm_edit_limit":
        tid = state_data.get("tariff_id")
        if text.strip() == "/empty":
            await db.update_tariff(tid, activation_limit=None)
        else:
            try:
                val = int(text.strip())
            except ValueError:
                await bot.send_message(chat_id, "Введите число:")
                return True
            await db.update_tariff(tid, activation_limit=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Лимит активаций обновлён ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Редактирование: медиа ─────────────────────────────
    if state == "adm_edit_media":
        tid = state_data.get("tariff_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_tariff(tid, media_url=val)
        clear_state(user_id)
        tariff = await db.get_tariff(tid)
        await bot.send_message(chat_id, "Медиа обновлено ✅")
        await _send_tariff_settings(bot, chat_id, tariff)
        return True

    # ── Профиль пользователя: ввод id ──────────────────────
    if state == "adm_profile_input":
        try:
            target_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой id пользователя:")
            return True
        profile = await db.user_profile(target_id)
        if not profile:
            await bot.send_message(
                chat_id, "Пользователь не найден. Убедитесь что он запускал бота.",
                keyboard=akb.admin_back_subscribers(),
            )
            clear_state(user_id)
            return True
        clear_state(user_id)
        await _send_user_profile(bot, chat_id, profile)
        return True

    # ── Поиск по ID: ввод id ────────────────────────────────
    if state == "adm_search_id_input":
        try:
            target_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой ID пользователя:")
            return True
        profile = await db.user_profile(target_id)
        if not profile:
            await bot.send_message(
                chat_id, "Пользователь не найден. Убедитесь что он запускал бота.",
                keyboard=akb.admin_back_subscribers(),
            )
            clear_state(user_id)
            return True
        clear_state(user_id)
        await _send_user_profile_with_logs(bot, chat_id, profile)
        return True

    # ── Выдать подписку: ввод user_id ──────────────────────
    if state == "adm_grant_user":
        try:
            target_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой user\\_ID:")
            return True
        tid = state_data.get("tariff_id")
        target = await db.get_user(target_id)
        if not target:
            await bot.send_message(
                chat_id,
                "Пользователь не найден. Он должен сначала написать /start боту.",
                keyboard=akb.admin_back_subscribers(),
            )
            clear_state(user_id)
            return True
        tariff = await db.get_tariff(tid)
        await db.grant_subscription(target_id, tid)
        clear_state(user_id)
        full_name = f"{target['first_name']} {target['last_name']}".strip()
        await bot.send_message(
            chat_id,
            f"✅ Подписка «{tariff['name']}» выдана пользователю "
            f"{full_name} (id={target_id})",
            keyboard=akb.admin_subscribers(),
        )
        return True

    # ── Обнулить подписку: ввод user_id ────────────────────
    if state == "adm_revoke_user":
        try:
            target_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой user\\_ID:")
            return True
        tid = state_data.get("tariff_id")
        tariff = await db.get_tariff(tid)
        await db.revoke_subscription(target_id, tid)
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"✅ Подписка «{tariff['name']}» обнулена у пользователя id={target_id}",
            keyboard=akb.admin_subscribers(),
        )
        return True

    # ── Передать подписку: ввод from user_id ───────────────
    if state == "adm_transfer_from":
        try:
            from_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой user\\_ID:")
            return True
        tid = state_data.get("tariff_id")
        set_state(user_id, "adm_transfer_to", tariff_id=tid, from_user_id=from_id)
        await bot.send_message(
            chat_id,
            f"Теперь отправьте user\\_ID пользователя, **которому** передать подписку:",
            keyboard=akb.admin_back_subscribers(),
        )
        return True

    if state == "adm_transfer_to":
        try:
            to_id = int(text.strip())
        except ValueError:
            await bot.send_message(chat_id, "Введите числовой user\\_ID:")
            return True
        tid = state_data.get("tariff_id")
        from_id = state_data.get("from_user_id")
        tariff = await db.get_tariff(tid)
        ok = await db.transfer_subscription(from_id, to_id, tid)
        clear_state(user_id)
        if ok:
            await bot.send_message(
                chat_id,
                f"✅ Подписка «{tariff['name']}» передана от id={from_id} к id={to_id}",
                keyboard=akb.admin_subscribers(),
            )
        else:
            await bot.send_message(
                chat_id,
                f"❌ У пользователя id={from_id} нет активной подписки «{tariff['name']}»",
                keyboard=akb.admin_subscribers(),
            )
        return True

    # ── Ответ на обратную связь ─────────────────────────────
    if state == "adm_reply_feedback":
        target_id = state_data.get("target_user_id")
        media_atts = [
            att for att in (attachments or [])
            if att.get("type") in ("image", "file", "video", "audio")
               and att.get("payload", {}).get("token")
        ]
        if text:
            reply_template = await db.get_bot_text("feedback_reply")
            await bot.send_message(
                target_id,
                reply_template.format(reply=text),
            )
        # Пересылаем медиавложения пользователю
        for att in media_atts:
            att_type = att.get("type", "file")
            token = att.get("payload", {}).get("token", "")
            if token:
                await bot.forward_attachment(target_id, att_type, token)
        # Возвращаем пользователю состояние диалога,
        # чтобы он мог продолжить переписку без повторного нажатия кнопки
        set_state(target_id, "waiting_feedback")
        clear_state(user_id)
        await bot.send_message(
            chat_id, "✅ Ответ отправлен пользователю.",
        )
        return True

    # ── Написать пользователю ──────────────────────────────
    if state == "adm_msg_user":
        target_id = state_data.get("target_user_id")
        await bot.send_message(target_id, text)
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"✅ Сообщение отправлено пользователю id={target_id}",
            keyboard=akb.admin_subscribers(),
        )
        return True

    # ══════════════════════════════════════════════════════════
    # ══  РАССЫЛКА  ════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════

    if state == "adm_broadcast":
        bc_group = state_data.get("bc_group", "all")
        bc_tariff_id = state_data.get("bc_tariff_id")
        set_state(user_id, "adm_broadcast_wait_tariff",
                  bc_group=bc_group, bc_tariff_id=bc_tariff_id, bc_text=text)
        tariffs = await db.list_tariffs()
        await bot.send_message(
            chat_id,
            "Выберите тариф, кнопка оформления которого будет добавлена к рассылке:",
            keyboard=akb.admin_broadcast_button_picker(tariffs),
        )
        return True

    if state == "adm_broadcast_wait_tariff":
        tariffs = await db.list_tariffs()
        await bot.send_message(
            chat_id,
            "Выберите тариф из списка или нажмите «Отправить без кнопки»:",
            keyboard=akb.admin_broadcast_button_picker(tariffs),
        )
        return True

    # ══════════════════════════════════════════════════════════
    # ══  ТЕКСТЫ БОТА  ═══════════════════════════════════════
    # ══════════════════════════════════════════════════════════

    if state == "adm_edit_bot_text":
        key = state_data.get("text_key")
        await db.set_bot_text(key, text.strip())
        clear_state(user_id)
        label = db.BOT_TEXT_LABELS.get(key, key)
        await bot.send_message(
            chat_id,
            f"✅ Текст «{label}» обновлён.",
            keyboard=akb.admin_bot_texts_list(db.BOT_TEXT_LABELS),
        )
        return True

    if state == "adm_edit_btn_text":
        key = state_data.get("text_key")
        await db.set_bot_text(key, text.strip())
        clear_state(user_id)
        label = db.BUTTON_TEXT_LABELS.get(key, key)
        await bot.send_message(
            chat_id,
            f"✅ Текст кнопки «{label}» обновлён на: {text.strip()}",
            keyboard=akb.admin_button_texts_list(db.BUTTON_TEXT_LABELS),
        )
        return True

    if state == "adm_edit_desc_text":
        key = state_data.get("text_key")
        await db.set_bot_text(key, text.strip())
        clear_state(user_id)
        label = db.DESC_TEXT_LABELS.get(key, key)
        await bot.send_message(
            chat_id,
            f"✅ Описание «{label}» обновлено.",
            keyboard=akb.admin_desc_texts_list(db.DESC_TEXT_LABELS),
        )
        return True

    # ══════════════════════════════════════════════════════════
    # ══  ПРОМОКОДЫ  ══════════════════════════════════════════
    # ══════════════════════════════════════════════════════════

    # ── Создание: ввод кода ──────────────────────────────────
    if state == "adm_promo_name":
        code = text.strip()
        existing = await db.get_promo_by_code(code)
        if existing:
            await bot.send_message(
                chat_id, f"❌ Промокод «{code}» уже существует. Введите другой:",
                keyboard=akb.admin_promo_back(),
            )
            return True
        promo_type = state_data.get("promo_type", "general")
        set_state(user_id, "adm_promo_discount", promo_code=code, promo_type=promo_type)
        await bot.send_message(
            chat_id,
            f"Промокод — **{code}**\n\n"
            "Отправьте боту процент скидки на покупку тарифов (1-100).",
            keyboard=akb.admin_promo_back(),
        )
        return True

    # ── Создание: ввод скидки ────────────────────────────────
    if state == "adm_promo_discount":
        try:
            pct = int(text.strip())
            if not 1 <= pct <= 100:
                raise ValueError
        except ValueError:
            await bot.send_message(chat_id, "Введите число от 1 до 100:")
            return True
        code = state_data.get("promo_code", "")
        promo_type = state_data.get("promo_type", "general")
        set_state(user_id, "adm_promo_max_act",
                  promo_code=code, promo_discount=pct, promo_type=promo_type)
        await bot.send_message(
            chat_id,
            f"**Создание промокода**\n\n"
            f"Промокод — **{code}**\n"
            f"Процент скидки — **{pct}%**\n"
            "Отправьте боту нужное количество активаций "
            "либо отправьте 0 для его безлимитного числа.",
            keyboard=akb.admin_promo_back(),
        )
        return True

    # ── Создание: ввод макс. активаций ───────────────────────
    if state == "adm_promo_max_act":
        try:
            max_act = int(text.strip())
            if max_act < 0:
                raise ValueError
        except ValueError:
            await bot.send_message(chat_id, "Введите неотрицательное число:")
            return True
        code = state_data.get("promo_code", "")
        pct = state_data.get("promo_discount", 0)
        promo_type = state_data.get("promo_type", "general")
        promo = await db.create_promo(code, pct, max_act, promo_type)
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"Промокод **{code}** создан ✅",
            keyboard=akb.admin_promo_created(promo["id"]),
        )
        return True

    # ── Создание группы промокодов ───────────────────────────
    if state == "adm_promo_group_name":
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"✅ Группа промокодов «{text.strip()}» создана.",
            keyboard=akb.admin_promo_menu(),
        )
        return True

    # ── Редактирование: макс. активаций ──────────────────────
    if state == "adm_promo_edit_max":
        try:
            val = int(text.strip())
            if val < 0:
                raise ValueError
        except ValueError:
            await bot.send_message(chat_id, "Введите неотрицательное число:")
            return True
        pid = state_data.get("promo_id")
        await db.update_promo(pid, max_activations=val)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await bot.send_message(chat_id, "✅ Кол-во активаций обновлено.")
        await _send_promo_detail(bot, chat_id, promo)
        return True

    # ── Редактирование: макс. на человека ────────────────────
    if state == "adm_promo_edit_per_user":
        try:
            val = int(text.strip())
            if val < 1:
                raise ValueError
        except ValueError:
            await bot.send_message(chat_id, "Введите число ≥ 1:")
            return True
        pid = state_data.get("promo_id")
        await db.update_promo(pid, max_per_user=val)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await bot.send_message(chat_id, "✅ Лимит на человека обновлён.")
        await _send_promo_detail(bot, chat_id, promo)
        return True

    # ── Редактирование: группа разрешённых ───────────────────
    if state == "adm_promo_edit_group":
        pid = state_data.get("promo_id")
        val = None if text.strip() == "/empty" else text.strip()
        await db.update_promo(pid, allowed_group=val)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await bot.send_message(chat_id, "✅ Группа разрешённых обновлена.")
        await _send_promo_detail(bot, chat_id, promo)
        return True

    # ── Редактирование: срок действия ────────────────────────
    if state == "adm_promo_edit_expiry":
        pid = state_data.get("promo_id")
        if text.strip() == "/empty":
            await db.update_promo(pid, expires_at=None)
        else:
            dt = _parse_datetime(text.strip())
            if not dt:
                await bot.send_message(chat_id, "Формат: ДД.ММ.ГГГГ ЧЧ:ММ")
                return True
            await db.update_promo(pid, expires_at=dt)
        clear_state(user_id)
        promo = await db.get_promo(pid)
        await bot.send_message(chat_id, "✅ Срок действия обновлён.")
        await _send_promo_detail(bot, chat_id, promo)
        return True

    # ── Способы оплаты: ввод названия ──────────────────────
    if state == "adm_pay_name":
        provider = state_data.get("provider", "")
        set_state(user_id, "adm_pay_shop_id", provider=provider, pay_name=text)
        await bot.send_message(
            chat_id,
            f"Название: **{text}**\n\nВведите Shop ID (идентификатор магазина):",
            keyboard=akb.admin_payment_cancel(),
        )
        return True

    # ── Способы оплаты: ввод Shop ID ──────────────────────
    if state == "adm_pay_shop_id":
        provider = state_data.get("provider", "")
        pay_name = state_data.get("pay_name", "")
        set_state(user_id, "adm_pay_secret", provider=provider,
                  pay_name=pay_name, shop_id=text)
        await bot.send_message(
            chat_id,
            f"Название: **{pay_name}**\nShop ID: **{text}**\n\n"
            "Введите секретный ключ (Secret Key):",
            keyboard=akb.admin_payment_cancel(),
        )
        return True

    # ── Способы оплаты: ввод Secret Key → создание ────────
    if state == "adm_pay_secret":
        provider = state_data.get("provider", "")
        pay_name = state_data.get("pay_name", "")
        shop_id = state_data.get("shop_id", "")
        method = await db.create_payment_method(pay_name, provider, shop_id, text)
        clear_state(user_id)
        await bot.send_message(
            chat_id,
            f"✅ Способ оплаты «{pay_name}» создан!\n"
            f"Провайдер: {provider}\nShop ID: {shop_id}",
            keyboard=akb.admin_payment_detail(method["id"], method["is_active"]),
        )
        return True

    return False


# ── Вспомогательные функции ───────────────────────────────────

async def _show_tariff_settings(reply_fn, tariff: dict):
    """Показывает экран настроек тарифа (через reply — edit)."""
    desc = tariff["description"] or "(пусто)"
    if len(desc) > 200:
        desc = desc[:200] + "…"
    await reply_fn(
        f"Настройка тарифа «{tariff['name']}»\n"
        f"Описание тарифа:\n{desc}",
        keyboard=akb.admin_tariff_settings(tariff["id"], tariff["is_active"]),
    )


async def _send_tariff_settings(bot: MaxBot, chat_id: int, tariff: dict):
    """Отправляет экран настроек тарифа (новое сообщение)."""
    desc = tariff["description"] or "(пусто)"
    if len(desc) > 200:
        desc = desc[:200] + "…"
    await bot.send_message(
        chat_id,
        f"Настройка тарифа «{tariff['name']}»\n"
        f"Описание тарифа:\n{desc}",
        keyboard=akb.admin_tariff_settings(tariff["id"], tariff["is_active"]),
    )


def _parse_nullable_float(text: str) -> float | None:
    if text.strip() == "/empty":
        return None
    try:
        return float(text.replace(",", ".").replace(" ", ""))
    except ValueError:
        return None


def _parse_datetime(text: str) -> datetime | None:
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


async def _send_user_profile(bot: MaxBot, chat_id: int, profile: dict):
    """Отправляет профиль пользователя."""
    uid = profile["user_id"]
    name = f"{profile['first_name']} {profile['last_name']}".strip() or "—"

    subs_lines = []
    for p in profile["purchases"]:
        date_str = p["purchased_at"].strftime("%d.%m.%Y") if p.get("purchased_at") else "?"
        subs_lines.append(f" • {p['tariff_name']} — {date_str}")
    subs_text = "\n".join(subs_lines) if subs_lines else " нет покупок"

    text = (
        f"User ID: **{uid}**\n"
        f"Имя пользователя: **{name}**\n\n"
        f"Подписки:\n{subs_text}\n\n"
        f"Кол-во оплат: **{profile['total_count']}**\n"
        f"Сумма оплат: **{profile['total_paid']:.0f}₽**\n"
        f"Средний чек: **{profile['avg_check']:.0f}₽**"
    )
    await bot.send_message(chat_id, text, keyboard=akb.admin_user_profile(uid))


async def _send_user_profile_with_logs(bot: MaxBot, chat_id: int, profile: dict):
    """Отправляет профиль пользователя с логами действий."""
    uid = profile["user_id"]
    name = f"{profile['first_name']} {profile['last_name']}".strip() or "—"

    # Активные тарифы
    active_tariffs = []
    for p in profile["purchases"]:
        if p.get("status") == "active":
            exp = p.get("expires_at")
            if exp is None or exp > datetime.now():
                active_tariffs.append(p["tariff_name"])
    active_str = ", ".join(active_tariffs) if active_tariffs else "нет"

    # Логи
    logs = await db.get_user_logs(uid)
    if logs:
        # Группируем по дате
        from collections import defaultdict
        by_date = defaultdict(list)
        for log in reversed(logs):
            date_str = log["created_at"].strftime("%d.%m.%Y")
            time_str = log["created_at"].strftime("%H:%M")
            by_date[date_str].append(f"{time_str}: {log['action']}")
        logs_lines = []
        for date, actions in by_date.items():
            logs_lines.append(f"\n**{date}**")
            for a in actions:
                logs_lines.append(a)
        logs_text = "\n".join(logs_lines)
    else:
        logs_text = "нет логов"

    text = (
        f"Пользователь: **{name}**\n"
        f"ID: **{uid}**\n"
        f"Активные тарифы: {active_str}\n\n"
        f"Логи:{logs_text}"
    )
    await bot.send_message(chat_id, text, keyboard=akb.admin_user_profile(uid))


async def _send_users_xlsx(bot: MaxBot, chat_id: int):
    """Генерирует и отправляет xlsx-таблицу пользователей."""
    from openpyxl import Workbook

    users = await db.all_users_with_purchases()
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    ws.append([
        "Имя Фамилия", "User ID", "Номер телефона",
        "Кол-во покупок", "Общая сумма оплат",
        "Купленные тарифы (✅ активен / ❌ нет)",
    ])
    for u in users:
        full_name = f"{u['first_name']} {u['last_name']}".strip()
        ws.append([
            full_name,
            u["user_id"],
            u.get("phone") or "",
            u["purchase_count"],
            float(u["total_paid"]),
            u["purchases"] or "",
        ])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        await bot.send_file(chat_id, tmp.name, "users.xlsx")
    finally:
        os.unlink(tmp.name)


async def _send_subscribers_xlsx(bot: MaxBot, chat_id: int):
    """Генерирует xlsx со списком активных подписчиков по тарифам."""
    from openpyxl import Workbook

    rows = await db.tariff_subscribers()
    wb = Workbook()
    # Группируем по тарифу — один лист на тариф
    sheets_created = set()
    default_removed = False

    for r in rows:
        tariff_name = r["tariff_name"] or "Без названия"
        # Имя листа (макс 31 символ для Excel)
        sheet_name = tariff_name[:31]
        if sheet_name not in sheets_created:
            if not default_removed:
                ws = wb.active
                ws.title = sheet_name
                default_removed = True
            else:
                ws = wb.create_sheet(title=sheet_name)
            ws.append([
                "Имя Фамилия", "User ID", "Телефон",
                "Дата покупки", "Время покупки", "Действует до", "Сумма оплаты",
            ])
            sheets_created.add(sheet_name)
        else:
            ws = wb[sheet_name]

        full_name = f"{r['first_name']} {r['last_name']}".strip()
        purchased_dt = r.get("purchased_at")
        purchased_date = purchased_dt.strftime("%d.%m.%Y") if purchased_dt else ""
        purchased_time = purchased_dt.strftime("%H:%M") if purchased_dt else ""
        expires = r["expires_at"].strftime("%d.%m.%Y") if r.get("expires_at") else "бессрочно"
        price_paid = r.get("price_paid")
        price_str = f"{price_paid:.0f}₽" if price_paid else "бесплатно"
        ws.append([full_name, r["user_id"], r.get("phone") or "", purchased_date, purchased_time, expires, price_str])

    if not rows:
        ws = wb.active
        ws.title = "Нет подписчиков"
        ws.append(["Активных подписчиков нет"])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        await bot.send_file(chat_id, tmp.name, "subscribers.xlsx")
    finally:
        os.unlink(tmp.name)


async def _send_expired_xlsx(bot: MaxBot, chat_id: int):
    """Генерирует xlsx со списком не продливших подписку."""
    from openpyxl import Workbook

    rows = await db.tariff_expired_subscribers()
    wb = Workbook()
    sheets_created = set()
    default_removed = False

    for r in rows:
        tariff_name = r["tariff_name"] or "Без названия"
        sheet_name = tariff_name[:31]
        if sheet_name not in sheets_created:
            if not default_removed:
                ws = wb.active
                ws.title = sheet_name
                default_removed = True
            else:
                ws = wb.create_sheet(title=sheet_name)
            ws.append([
                "Имя Фамилия", "User ID", "Телефон",
                "Дата покупки", "Истекла",
            ])
            sheets_created.add(sheet_name)
        else:
            ws = wb[sheet_name]

        full_name = f"{r['first_name']} {r['last_name']}".strip()
        purchased = r["purchased_at"].strftime("%d.%m.%Y") if r.get("purchased_at") else ""
        expired = r["expires_at"].strftime("%d.%m.%Y") if r.get("expires_at") else ""
        ws.append([full_name, r["user_id"], r.get("phone") or "", purchased, expired])

    if not rows:
        ws = wb.active
        ws.title = "Нет данных"
        ws.append(["Не продливших подписку нет"])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        await bot.send_file(chat_id, tmp.name, "expired.xlsx")
    finally:
        os.unlink(tmp.name)


# ── Промокоды: вспомогательные ────────────────────────────────

async def _show_promo_detail(reply_fn, promo: dict):
    """Показывает детали промокода (через reply — edit)."""
    max_act = promo["max_activations"]
    max_act_str = "безлимит" if max_act == 0 else str(max_act)
    expires = promo["expires_at"].strftime("%d.%m.%Y %H:%M") if promo.get("expires_at") else "не ограничен"
    used = await db.count_promo_activations(promo["id"])
    text = (
        f"Промокод — **{promo['code']}**\n"
        f"Процент скидки — **{promo['discount_percent']}%**\n"
        f"Максимум активаций: **{max_act_str}**\n"
        f"Максимум активаций одним человеком — **{promo['max_per_user']}**\n"
        f"Срок действия: **{expires}**\n"
        f"Использовано: **{used}**"
    )
    await reply_fn(text, keyboard=akb.admin_promo_detail(promo["id"]))


async def _send_promo_detail(bot: MaxBot, chat_id: int, promo: dict):
    """Отправляет детали промокода (новое сообщение)."""
    max_act = promo["max_activations"]
    max_act_str = "безлимит" if max_act == 0 else str(max_act)
    expires = promo["expires_at"].strftime("%d.%m.%Y %H:%M") if promo.get("expires_at") else "не ограничен"
    used = await db.count_promo_activations(promo["id"])
    text = (
        f"Промокод — **{promo['code']}**\n"
        f"Процент скидки — **{promo['discount_percent']}%**\n"
        f"Максимум активаций: **{max_act_str}**\n"
        f"Максимум активаций одним человеком — **{promo['max_per_user']}**\n"
        f"Срок действия: **{expires}**\n"
        f"Использовано: **{used}**"
    )
    await bot.send_message(chat_id, text, keyboard=akb.admin_promo_detail(promo["id"]))


async def _send_promo_activations_xlsx(bot: MaxBot, chat_id: int, promo_id: int):
    """Генерирует и отправляет xlsx-таблицу активаций промокода."""
    from openpyxl import Workbook

    activations = await db.get_promo_activations(promo_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Активации"
    ws.append(["ID", "Имя Фамилия", "Оплатил"])
    for a in activations:
        full_name = f"{a.get('first_name', '')} {a.get('last_name', '')}".strip()
        paid_str = "Да" if a.get("paid") else "Нет"
        ws.append([a["user_id"], full_name, paid_str])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    try:
        await bot.send_file(chat_id, tmp.name, "promo_activations.xlsx")
    finally:
        os.unlink(tmp.name)


# ── Рассылка / контакты: вспомогательные ─────────────────────

def _contact_request_kb() -> dict:
    """Клавиатура с кнопкой запроса контакта."""
    return {
        "type": "inline_keyboard",
        "payload": {"buttons": [
            [{"type": "request_contact", "text": "📱 Отправить номер телефона"}],
        ]},
    }